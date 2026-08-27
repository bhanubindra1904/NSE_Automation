"""Download recent NSE disclosures and build one audited workbook per company.

The browser collection is intentionally slow and visible. PDF classification is
action based: a document is only marked as actual capital or debt when the text
states that shares/debt were allotted, issued, subscribed, raised, or drawn.
Ratings, proposals, letterhead identifiers, and historical references remain
financially relevant without being counted as completed financing events.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus, urljoin, urlparse

try:
    import fitz  # PyMuPDF
except ImportError:  # pragma: no cover
    fitz = None

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover
    PdfReader = None

try:
    import pandas as pd
except ImportError:  # pragma: no cover
    pd = None

try:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright
except ImportError:  # pragma: no cover
    PlaywrightTimeoutError = Exception
    sync_playwright = None


NSE_HOME = "https://www.nseindia.com/"
NSE_ANNOUNCEMENTS_PAGE = "https://www.nseindia.com/companies-listing/corporate-filings-announcements"

DEFAULT_SYMBOL = "DELHIVERY"
DEFAULT_COMPANY_PREFIX = "DELHIVERY"
DEFAULT_COMPANY_NAME = "Delhivery Limited"

CURRENCY_TOKEN = r"(?<![A-Za-z])(?:\u20b9|rs\s*\.?|inr|usd)(?![A-Za-z])"
NUMBER_TOKEN = r"[0-9][0-9,]*(?:\.[0-9]+)?"
AMOUNT_UNIT_TOKEN = r"(?:crores?|cr|lakhs?|lacs?|mn|million|billion)"
AMOUNT_PATTERN = re.compile(
    rf"{CURRENCY_TOKEN}\s*{NUMBER_TOKEN}\s*(?:{AMOUNT_UNIT_TOKEN})?"
    rf"|{NUMBER_TOKEN}\s*{AMOUNT_UNIT_TOKEN}",
    re.IGNORECASE,
)
PERCENT_TOKEN = r"(?:%|\u00b0\s*/\s*[o0]|[o0]\s*/\s*[o0])"
INDIAN_GROUPED_NUMBER_PATTERN = re.compile(
    r"(?<![0-9])([0-9]{1,2}\s*(?:,\s*|\s+)[0-9](?:\s*[0-9])\s*(?:,\s*|\s+)[0-9]{3})(?![0-9])"
)

DISCLOSURE_LINK_SCRIPT = r"""
({symbol, companyPrefix, companyName, requireCompany}) => {
  const clean = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    const rect = el.getBoundingClientRect();
    return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
  };
  const rowContext = (el) => el.closest('tr, [role="row"], .ag-row, .card, li, section, article, div') || el;

  document.querySelectorAll('[data-nse-disclosure-download]').forEach((el) => {
    el.removeAttribute('data-nse-disclosure-download');
  });

  const symbolU = (symbol || '').toUpperCase();
  const prefixU = (companyPrefix || '').toUpperCase();
  const companyU = (companyName || '').toUpperCase();
  const candidates = [];

  const clickableElements = Array.from(document.querySelectorAll('a[href], button, [role="button"]'))
    .filter(visible);

  for (const el of clickableElements) {
    const href = el.href || el.getAttribute('href') || '';
    const text = clean(el.innerText || el.textContent || el.getAttribute('aria-label') || el.getAttribute('title') || '');
    const ctxEl = rowContext(el);
    const context = clean(ctxEl.innerText || ctxEl.textContent || text);
    const combined = `${href} ${text} ${context}`.toUpperCase();
    const hrefLower = href.toLowerCase();

    const looksLikeAttachment =
      hrefLower.includes('.pdf') ||
      hrefLower.includes('.xml') ||
      hrefLower.includes('/corporate/') ||
      hrefLower.includes('attachment') ||
      hrefLower.includes('attchmnt') ||
      combined.includes('PDF') ||
      combined.includes('ATTACHMENT') ||
      combined.includes('DOWNLOAD');

    const companyRelated =
      combined.includes(symbolU) ||
      combined.includes(prefixU + ' ') ||
      combined.includes(prefixU + '_') ||
      combined.includes(prefixU + '-') ||
      (companyU && combined.includes(companyU));

    if (!looksLikeAttachment) continue;
    if (requireCompany && !companyRelated) continue;

    const selectorValue = String(candidates.length);
    el.setAttribute('data-nse-disclosure-download', selectorValue);
    candidates.push({
      index: candidates.length,
      href,
      text,
      context,
      companyRelated,
      selector: `[data-nse-disclosure-download="${selectorValue}"]`
    });
  }

  return candidates.slice(0, 150);
}
"""

SCROLL_PAGE_SCRIPT = r"""
({pixels}) => {
  const before = window.scrollY || document.documentElement.scrollTop || 0;
  const amount = Math.max(200, Number(pixels || 300));
  window.scrollBy(0, amount);
  const after = window.scrollY || document.documentElement.scrollTop || 0;
  return {moved: after > before + 2, before, after};
}
"""

CLICK_MORE_RESULTS_SCRIPT = r"""
() => {
  const clean = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const style = window.getComputedStyle(el);
    const rect = el.getBoundingClientRect();
    return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
  };
  const candidates = Array.from(document.querySelectorAll('button, a, [role="button"], li, span'))
    .filter(visible)
    .map((el) => ({el, text: clean(el.innerText || el.textContent || '').toUpperCase(), aria: clean(el.getAttribute('aria-label') || '').toUpperCase()}));
  const item = candidates.find((x) =>
    ['LOAD MORE', 'SHOW MORE', 'MORE RESULTS', 'NEXT'].includes(x.text) ||
    x.aria.includes('LOAD MORE') || x.aria.includes('NEXT')
  );
  if (!item) return false;
  const target = item.el.closest('button, a, [role="button"], li') || item.el;
  target.scrollIntoView({block: 'center'});
  target.click();
  return true;
}
"""

BODY_TEXT_SCRIPT = "() => document.body ? (document.body.innerText || document.body.textContent || '') : ''"


@dataclass
class DownloadResult:
    downloaded_path: Path | None
    actual_pdf_url: str
    status: str
    error: str


def clean_text(value: Any) -> str:
    text = str(value or "").replace("\x00", " ").replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


EXCEL_ILLEGAL_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean_excel_value(value: Any) -> Any:
    """Remove values that can make openpyxl fail halfway through a workbook."""
    if value is None or isinstance(value, (int, float, bool, datetime)):
        return value
    return EXCEL_ILLEGAL_CHARACTERS.sub(" ", str(value))[:32700]


def sanitize_records(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {clean_excel_value(key): clean_excel_value(value) for key, value in row.items()}
        for row in rows
    ]


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalized_text_sha256(text: str) -> str:
    """Hash semantic text so binary-different copies of one PDF still dedupe."""
    normalized = re.sub(r"[^a-z0-9]+", " ", clean_text(text).casefold()).strip()
    if not normalized:
        return ""
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def disclosure_hash_key(disclosure_date: Any, digest: Any) -> str:
    """Scope content dedupe to one official NSE disclosure date."""
    clean_digest = clean_text(digest)
    if not clean_digest:
        return ""
    return "|".join([clean_text(disclosure_date), clean_digest])


def safe_filename(value: str, fallback: str = "nse_disclosure.pdf") -> str:
    cleaned = clean_text(value)
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", cleaned)
    cleaned = re.sub(r"\s+", "_", cleaned).strip("._ ")[:180]
    return cleaned or fallback


def require_dependencies() -> None:
    missing = []
    if sync_playwright is None:
        missing.append("playwright")
    if fitz is None and PdfReader is None:
        missing.append("pymupdf (or pypdf)")
    if pd is None:
        missing.append("pandas/openpyxl")
    if missing:
        print("Missing package(s): " + ", ".join(missing))
        print("Install with: pip install -r requirements.txt")
        print("Then run: python -m playwright install chromium")
        sys.exit(2)


def announcements_url(args: argparse.Namespace) -> str:
    return f"{NSE_ANNOUNCEMENTS_PAGE}?symbol={quote_plus(args.symbol)}&tabIndex=equity"


def parse_disclosure_date(text: str) -> datetime | None:
    text = clean_text(text)

    # NSE announcement rows end with the exchange timestamp. Prefer that over
    # dates embedded in attachment names or over event dates such as record dates.
    timestamp_patterns: list[tuple[str, list[str]]] = [
        (r"\b(\d{1,2}-[A-Za-z]{3}-20\d{2})\s+\d{1,2}:\d{2}(?::\d{2})?\b", ["%d-%b-%Y"]),
        (r"\b(\d{1,2}\s+[A-Za-z]{3,9}\s+20\d{2})\s+(?:at\s+)?\d{1,2}:\d{2}(?::\d{2})?\b", ["%d %b %Y", "%d %B %Y"]),
        (r"\b(20\d{2}-\d{2}-\d{2})[ T]\d{1,2}:\d{2}(?::\d{2})?\b", ["%Y-%m-%d"]),
        (r"\b(\d{1,2}[-/]\d{1,2}[-/]20\d{2})\s+\d{1,2}:\d{2}(?::\d{2})?\b", ["%d-%m-%Y", "%d/%m/%Y", "%m-%d-%Y", "%m/%d/%Y"]),
    ]
    timestamp_candidates: list[datetime] = []
    for pattern, formats in timestamp_patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            for fmt in formats:
                try:
                    timestamp_candidates.append(datetime.strptime(match.group(1), fmt))
                    break
                except ValueError:
                    pass
    if timestamp_candidates:
        return timestamp_candidates[-1]

    leading_iso = re.match(r"^(20\d{2}-\d{2}-\d{2})(?:\b|_)", text)
    if leading_iso:
        try:
            return datetime.strptime(leading_iso.group(1), "%Y-%m-%d")
        except ValueError:
            pass

    candidates: list[tuple[str, list[str]]] = []

    for match in re.finditer(r"_(\d{2})(\d{2})(20\d{2})\d{6}_", text):
        candidates.append((f"{match.group(1)}-{match.group(2)}-{match.group(3)}", ["%d-%m-%Y"]))

    for match in re.finditer(r"\b\d{1,2}[-/]\d{1,2}[-/]20\d{2}\b", text):
        candidates.append((match.group(0), ["%d-%m-%Y", "%d/%m/%Y", "%m-%d-%Y", "%m/%d/%Y"]))
    for match in re.finditer(r"\b\d{1,2}-[A-Za-z]{3}-20\d{2}\b", text):
        candidates.append((match.group(0), ["%d-%b-%Y"]))
    for match in re.finditer(r"\b\d{1,2}\s+[A-Za-z]{3,9}\s+20\d{2}\b", text):
        candidates.append((match.group(0), ["%d %b %Y", "%d %B %Y"]))
    for match in re.finditer(r"\b[A-Za-z]{3,9}\s+\d{1,2},\s+20\d{2}\b", text):
        candidates.append((match.group(0), ["%b %d, %Y", "%B %d, %Y"]))
    for match in re.finditer(r"\b20\d{2}-\d{2}-\d{2}\b", text):
        candidates.append((match.group(0), ["%Y-%m-%d"]))

    for value, formats in candidates:
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
    return None


def open_page(page, url: str, args: argparse.Namespace, label: str) -> None:
    print(f"[NSE] Opening {label}: {url}")
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=args.page_timeout_seconds * 1000 if args.page_timeout_seconds > 0 else 0)
    except Exception as exc:
        print(f"[NSE] Navigation warning for {label}: {exc}")
        print("      If the browser is open, let it finish loading manually.")
    if args.manual_ready:
        print(f"[NSE] Manual mode: wait until {label} is visible, then press ENTER in PowerShell.")
        input("Press ENTER when ready...")
        return
    try:
        page.wait_for_function(
            "() => location.href !== 'about:blank' && document.body && (document.body.innerText || '').trim().length > 50",
            timeout=args.page_timeout_seconds * 1000 if args.page_timeout_seconds > 0 else 0,
        )
    except PlaywrightTimeoutError:
        print(f"[NSE] {label} was not detected automatically.")
        input("Press ENTER when ready...")


def wait_for_page_shell_loaded(page, args: argparse.Namespace, label: str = "page") -> None:
    print(f"[NSE] Waiting for {label} shell to finish loading...")
    for state in ["domcontentloaded", "load"]:
        try:
            page.wait_for_load_state(state, timeout=min(args.page_timeout_seconds, 30) * 1000 if args.page_timeout_seconds > 0 else 0)
        except Exception:
            pass
    try:
        page.wait_for_function(
            "() => document.readyState === 'complete' && document.body && (document.body.innerText || '').trim().length > 80",
            timeout=min(args.page_timeout_seconds, 30) * 1000 if args.page_timeout_seconds > 0 else 0,
        )
    except Exception:
        pass
    page.wait_for_timeout(int(args.after_load_pause_seconds * 1000))


def page_has_loading_indicator(page) -> bool:
    try:
        return bool(page.evaluate(
            """() => {
                const clean = (v) => (v || '').replace(/\\s+/g, ' ').trim().toUpperCase();
                const visible = (el) => {
                    if (!el) return false;
                    const style = window.getComputedStyle(el);
                    const rect = el.getBoundingClientRect();
                    return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
                };
                const text = clean(document.body ? document.body.innerText : '');
                if (text.includes('LOADING') || text.includes('PLEASE WAIT')) return true;
                return Array.from(document.querySelectorAll('.loader, .spinner, [class*="loader"], [class*="spinner"], [aria-busy="true"]')).some(visible);
            }"""
        ))
    except Exception:
        return False


def wait_until_not_loading(page, args: argparse.Namespace, label: str = "NSE content") -> None:
    import time
    deadline = time.time() + max(5, args.list_ready_timeout_seconds)
    while time.time() < deadline:
        if not page_has_loading_indicator(page):
            page.wait_for_timeout(int(args.after_load_pause_seconds * 1000))
            return
        print(f"[NSE] {label} still loading; waiting before scrolling...")
        page.wait_for_timeout(int(args.scroll_pause_seconds * 1000))


def wait_for_disclosure_rows_loaded(page, args: argparse.Namespace, get_candidates) -> None:
    print("[NSE] Waiting for disclosure rows/attachments before scrolling...")
    import time
    deadline = time.time() + max(10, args.list_ready_timeout_seconds)
    last_preview = ""

    while time.time() < deadline:
        wait_until_not_loading(page, args, "Disclosure list")
        # NSE page is already opened with ?symbol=<company>, so every visible attachment
        # belongs to that company. Do NOT require symbol/company text in the PDF filename.
        # Use --require-company-text-in-attachment only to revert to the old strict behavior.
        candidates = get_candidates(require_company=bool(getattr(args, "require_company_text_in_attachment", False)))

        if candidates:
            print(f"[NSE] Rows ready: {len(candidates)} visible attachment/link candidate(s).")
            page.wait_for_timeout(int(args.after_load_pause_seconds * 1000))
            return

        try:
            body = clean_text(page.evaluate(BODY_TEXT_SCRIPT))
        except Exception:
            body = ""

        upper = body.upper()
        has_company = args.symbol.upper() in upper or args.company_name.upper() in upper
        has_table_signal = any(word in upper for word in ["ATTACHMENT", "READ MORE", "SUBJECT", "ANNOUNCEMENT", "DATE"])
        if has_company and has_table_signal:
            print("[NSE] Disclosure list text is visible; waiting one extra pause before collecting.")
            page.wait_for_timeout(int(args.after_load_pause_seconds * 1000))
            return

        preview = body[:180]
        if preview and preview != last_preview:
            print(f"    waiting for rows... page preview: {preview!r}")
            last_preview = preview
        page.wait_for_timeout(int(args.scroll_pause_seconds * 1000))

    print("[NSE] Disclosure rows were not detected automatically.")
    input("Press ENTER when rows are visible...")


def wait_after_scroll_or_click(page, args: argparse.Namespace, get_candidates, previous_visible_count: int) -> None:
    import time
    deadline = time.time() + max(5, args.row_update_timeout_seconds)
    while time.time() < deadline:
        wait_until_not_loading(page, args, "Rows after scroll/click")
        # NSE page is already opened with ?symbol=<company>, so every visible attachment
        # belongs to that company. Do NOT require symbol/company text in the PDF filename.
        # Use --require-company-text-in-attachment only to revert to the old strict behavior.
        candidates = get_candidates(require_company=bool(getattr(args, "require_company_text_in_attachment", False)))
        if len(candidates) != previous_visible_count:
            print(f"[NSE] Row update detected: visible candidates {previous_visible_count} -> {len(candidates)}")
            page.wait_for_timeout(int(args.after_load_pause_seconds * 1000))
            return
        page.wait_for_timeout(int(args.scroll_pause_seconds * 1000))
    print("[NSE] No visible row-count change after scroll/click; continuing slowly.")
    page.wait_for_timeout(int(args.after_load_pause_seconds * 1000))


def collect_disclosure_candidates(page, args: argparse.Namespace) -> list[dict[str, Any]]:
    print(f"[NSE] Collecting visible {args.company_name} disclosure PDFs from the last {args.lookback_days} days.")
    cutoff = datetime.now() - timedelta(days=args.lookback_days)
    latest_allowed = datetime.now() + timedelta(days=1)
    seen: dict[str, dict[str, Any]] = {}
    stable_rounds = 0
    previous_count = -1
    older_seen_rounds = 0

    def get_candidates(require_company: bool = False) -> list[dict[str, Any]]:
        try:
            return page.evaluate(
                DISCLOSURE_LINK_SCRIPT,
                {"symbol": args.symbol, "companyPrefix": args.company_prefix, "companyName": args.company_name, "requireCompany": require_company},
            )
        except Exception as exc:
            print(f"[NSE] Candidate extraction warning: {exc}")
            return []

    wait_for_page_shell_loaded(page, args, f"{args.company_name} announcements page")
    wait_for_disclosure_rows_loaded(page, args, get_candidates)

    for check_no in range(1, args.max_scroll_checks + 1):
        wait_until_not_loading(page, args, "Disclosure list")
        # NSE page is already opened with ?symbol=<company>, so every visible attachment
        # belongs to that company. Do NOT require symbol/company text in the PDF filename.
        # Use --require-company-text-in-attachment only to revert to the old strict behavior.
        candidates = get_candidates(require_company=bool(getattr(args, "require_company_text_in_attachment", False)))
        visible_count_before_action = len(candidates)

        dated_this_round = 0
        older_this_round = 0
        for candidate in candidates:
            context = clean_text(candidate.get("context"))
            href = clean_text(candidate.get("href"))
            text = clean_text(candidate.get("text"))
            date_value = parse_disclosure_date(context)
            if not date_value:
                date_value = parse_disclosure_date(" ".join([href, text]))
            candidate["parsed_date"] = date_value.strftime("%Y-%m-%d") if date_value else ""
            candidate["within_lookback"] = bool(date_value and cutoff.date() <= date_value.date() <= latest_allowed.date())

            if date_value:
                dated_this_round += 1
                if date_value.date() < cutoff.date():
                    older_this_round += 1

            include = candidate["within_lookback"] or (not date_value and args.include_undated_visible)
            if not include:
                continue

            # Keep raw candidates for now; final dedupe happens after resolving Actual PDF URL.
            key = clean_text(candidate.get("href")) or clean_text(candidate.get("context"))[:240]
            if key and key not in seen:
                seen[key] = candidate

        print(f"    check {check_no}: {len(seen)} disclosure link(s) kept for the {args.lookback_days}-day window ({dated_this_round} dated visible, {older_this_round} older visible)")

        stable_rounds = stable_rounds + 1 if len(seen) == previous_count else 0
        previous_count = len(seen)
        older_seen_rounds = older_seen_rounds + 1 if older_this_round and dated_this_round and seen else 0

        if seen and (stable_rounds >= args.stable_rounds or older_seen_rounds >= 2):
            print("[NSE] Disclosure list appears complete for the lookback window; stopping collection.")
            break

        wait_until_not_loading(page, args, "Disclosure list before load-more/scroll")
        try:
            clicked_more = bool(page.evaluate(CLICK_MORE_RESULTS_SCRIPT))
        except Exception:
            clicked_more = False
        if clicked_more:
            print("[NSE] Clicked load-more/next; waiting for new rows.")
            wait_after_scroll_or_click(page, args, get_candidates, visible_count_before_action)
            continue

        try:
            moved_info = page.evaluate(SCROLL_PAGE_SCRIPT, {"pixels": args.scroll_pixels})
            moved = bool(moved_info.get("moved"))
            if moved:
                print(f"[NSE] Scrolled slowly from {moved_info.get('before')} to {moved_info.get('after')}; waiting for rows.")
                wait_after_scroll_or_click(page, args, get_candidates, visible_count_before_action)
        except Exception:
            moved = False

        if not moved and seen:
            stable_rounds += 1
        page.wait_for_timeout(int(args.scroll_pause_seconds * 1000))

    rows = list(seen.values())
    rows.sort(key=lambda item: item.get("parsed_date") or "9999-99-99", reverse=True)
    return rows[: args.max_documents] if args.max_documents > 0 else rows


def extract_attachment_url_from_bytes(body: bytes) -> str:
    try:
        text = body.decode("utf-8", errors="ignore")
    except Exception:
        text = body.decode("latin-1", errors="ignore")
    match = re.search(r"<[^>]*AttachmentURL[^>]*>\s*(https?://[^<\s]+\.pdf)\s*</", text, re.IGNORECASE)
    if match:
        return clean_text(match.group(1))
    match = re.search(r"https?://[^\s<>'\"]+\.pdf", text, re.IGNORECASE)
    return clean_text(match.group(0)) if match else ""


def filename_from_candidate(candidate: dict[str, Any], args: argparse.Namespace, actual_pdf_url: str = "") -> str:
    href = clean_text(actual_pdf_url or candidate.get("href"))
    basename = Path(urlparse(href).path).name if href else ""
    date_prefix = clean_text(candidate.get("parsed_date"))
    name = basename or clean_text(candidate.get("text")) or f"{args.symbol}_NSE_disclosure.pdf"
    filename = safe_filename(name, f"{args.symbol}_NSE_disclosure.pdf")
    if not Path(filename).suffix:
        filename += ".pdf"
    if filename.lower().endswith((".xml", ".xbrl")):
        filename = filename.rsplit(".", 1)[0] + ".pdf"
    if date_prefix and not filename.startswith(date_prefix):
        filename = f"{date_prefix}_{filename}"
    return filename


def fetch_url(context, url: str, args: argparse.Namespace) -> tuple[int, bytes]:
    response = context.request.get(
        urljoin(NSE_HOME, url),
        headers={
            "Accept": "application/pdf,application/xml,text/xml,text/html,*/*",
            "Accept-Encoding": "identity",
            "Cache-Control": "no-cache",
            "Referer": NSE_ANNOUNCEMENTS_PAGE,
            "User-Agent": args.user_agent,
        },
        timeout=args.download_timeout_seconds * 1000 if args.download_timeout_seconds > 0 else 0,
    )
    return response.status, response.body()


def write_pdf_bytes(dest: Path, body: bytes) -> bool:
    if not body or len(body) < 500:
        return False
    pdf_start = body.find(b"%PDF")
    if pdf_start < 0 or pdf_start > 1024:
        return False
    if pdf_start:
        body = body[pdf_start:]
    tmp = dest.with_suffix(dest.suffix + ".part")
    if tmp.exists():
        tmp.unlink()
    tmp.write_bytes(body)
    valid_pdf = False
    if fitz is not None:
        try:
            doc = fitz.open(str(tmp))
            page_count = doc.page_count
            doc.close()
            valid_pdf = page_count > 0
        except Exception:
            valid_pdf = False
    if not valid_pdf and PdfReader is not None:
        try:
            valid_pdf = len(PdfReader(str(tmp), strict=False).pages) > 0
        except Exception:
            valid_pdf = False
    if not valid_pdf:
        try:
            tmp.unlink()
        except Exception:
            pass
        return False
    if dest.exists():
        dest.unlink()
    tmp.replace(dest)
    return True


def download_candidate(context, candidate: dict[str, Any], downloads_dir: Path, args: argparse.Namespace) -> DownloadResult:
    downloads_dir.mkdir(parents=True, exist_ok=True)
    href = clean_text(candidate.get("href"))

    last_error = "Unknown download error"
    actual_pdf_url = href
    attempts = max(1, int(getattr(args, "download_retries", 2)) + 1)

    for attempt in range(1, attempts + 1):
        try:
            status, body = fetch_url(context, href, args)
            if status != 200:
                last_error = f"HTTP {status}"
                raise RuntimeError(last_error)

            actual_pdf_url = href
            if not body.startswith(b"%PDF"):
                actual_pdf_url = extract_attachment_url_from_bytes(body)
                if not actual_pdf_url:
                    return DownloadResult(None, href, "Failed", "No PDF bytes and no AttachmentURL found")
                status, body = fetch_url(context, actual_pdf_url, args)
                if status != 200:
                    last_error = f"AttachmentURL HTTP {status}"
                    raise RuntimeError(last_error)

            dest = downloads_dir / filename_from_candidate(candidate, args, actual_pdf_url=actual_pdf_url)
            if write_pdf_bytes(dest, body):
                status_text = "Downloaded" if attempt == 1 else f"Downloaded after {attempt} attempts"
                return DownloadResult(dest, actual_pdf_url, status_text, "")
            last_error = "PDF invalid/incomplete"
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"

        if attempt < attempts:
            print(f"[NSE] Download attempt {attempt}/{attempts} failed; retrying: {last_error}")
            time.sleep(max(0, float(getattr(args, "retry_pause_seconds", 2.0))))

    return DownloadResult(None, actual_pdf_url, "Failed", last_error)


def clean_pdf_text(value: str) -> str:
    return clean_text(value)


def extract_pdf_text(path: Path) -> tuple[str, list[dict[str, Any]], str]:
    pages: list[dict[str, Any]] = []
    if fitz is not None:
        try:
            doc = fitz.open(str(path))
        except Exception as exc:
            return "", [], f"Open failed: {type(exc).__name__}: {exc}"
        try:
            for idx, page in enumerate(doc, start=1):
                try:
                    page_text = clean_pdf_text(page.get_text("text") or "")
                except Exception as exc:
                    page_text = ""
                    pages.append({"Page": idx, "Text": "", "Parse Note": f"Page extraction failed: {type(exc).__name__}: {exc}"})
                    continue
                pages.append({"Page": idx, "Text": page_text, "Parse Note": ""})
        finally:
            doc.close()
    else:
        try:
            reader = PdfReader(str(path))
            for idx, page in enumerate(reader.pages, start=1):
                try:
                    page_text = clean_pdf_text(page.extract_text() or "")
                except Exception as exc:
                    page_text = ""
                    pages.append({"Page": idx, "Text": "", "Parse Note": f"Page extraction failed: {type(exc).__name__}: {exc}"})
                    continue
                pages.append({"Page": idx, "Text": page_text, "Parse Note": ""})
        except Exception as exc:
            return "", [], f"Open failed: {type(exc).__name__}: {exc}"

    full_text = "\n".join(row["Text"] for row in pages)
    if not clean_text(full_text):
        return "", pages, "No selectable text found; PDF may be scanned/image-only"
    return full_text, pages, "Parsed"


def context_window(text: str, start: int, end: int, chars: int = 340) -> str:
    return clean_text(text[max(0, start - chars): min(len(text), end + chars)])


def first_snippet(text: str, pattern: str, chars: int = 340) -> str:
    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    return context_window(text, match.start(), match.end(), chars)


def extract_subject(text: str, fallback: str = "") -> str:
    for pattern in [r"\bSub\s*:\s*(.*?)(?:Dear\s|Pursuant\s|Please\s|In continuation|$)", r"\bSubject\s*:\s*(.*?)(?:Dear\s|Pursuant\s|Please\s|In continuation|$)"]:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            value = clean_text(match.group(1))
            if value:
                return value[:300]
    return clean_text(fallback)[:300]


def extract_amounts(text: str) -> list[str]:
    out = []
    seen = set()
    for match in AMOUNT_PATTERN.finditer(text or ""):
        amount = clean_text(match.group(0))
        if amount.lower() not in seen:
            seen.add(amount.lower())
            out.append(amount)
    return out


def _amount_occurrences(text: str) -> list[tuple[str, int, int]]:
    return [
        (clean_text(match.group(0)), match.start(), match.end())
        for match in AMOUNT_PATTERN.finditer(text or "")
    ]


def _amount_context_label(text: str, amount_start: int) -> str:
    prefix = text[max(0, amount_start - 140):amount_start]
    label_patterns = [
        ("Money realized", r"money reali[sz]ed"),
        ("Exercise price", r"exercise price|pricing formula"),
        ("Paid-up capital", r"paid\s*-?\s*up(?: equity)? share capital"),
        ("Tax demand", r"tax demand"),
        ("Penalty", r"penalty"),
        ("Rated amount", r"rated amount|size of issue"),
        ("Cash consideration", r"cash consideration"),
        ("Cost of acquisition/subscription", r"cost of (?:acquisition|subscription)"),
        ("Issue price", r"issue price"),
        ("Face value", r"face value"),
        ("Consideration", r"consideration"),
    ]
    best_label = ""
    best_position = -1
    for label, pattern in label_patterns:
        for match in re.finditer(pattern, prefix, re.IGNORECASE):
            if match.start() > best_position:
                best_label = label
                best_position = match.start()
    return best_label


def _format_labeled_amounts(text: str) -> list[str]:
    values: list[str] = []
    for amount, start, _ in _amount_occurrences(text):
        label = _amount_context_label(text, start)
        display_amount = _amount_with_contextual_unit(amount, label, text)
        values.append(f"{label}: {display_amount}" if label else display_amount)
    return _unique_values(values)


def _amount_with_contextual_unit(amount: str, label: str, text: str) -> str:
    if label != "Rated amount" or re.search(AMOUNT_UNIT_TOKEN, amount, re.IGNORECASE):
        return amount
    unit_match = re.search(
        r"(?:rated amount|amount\s*/\s*size of issue).{0,60}\(\s*in\s+([a-z]+)\s*\)",
        text,
        re.IGNORECASE,
    )
    return f"{amount} {clean_text(unit_match.group(1)).lower()}" if unit_match else amount


def _format_indian_integer(value: str) -> str:
    digits = re.sub(r"[^0-9]", "", value or "")
    if len(digits) <= 3:
        return digits
    tail = digits[-3:]
    head = digits[:-3]
    groups: list[str] = []
    while head:
        groups.insert(0, head[-2:])
        head = head[:-2]
    return ",".join(groups + [tail])


def _extract_tax_metrics(text: str) -> list[str]:
    metrics: list[str] = []
    whole_number_values: list[int] = []
    for label, pattern in [
        ("Tax demand", r"tax demand"),
        ("Penalty", r"penalty"),
    ]:
        match = re.search(
            rf"\b{pattern}\b(?:\s+of)?\s*({CURRENCY_TOKEN}\s*{NUMBER_TOKEN}(?:\s*{AMOUNT_UNIT_TOKEN})?)",
            text,
            re.IGNORECASE,
        )
        if match:
            amount = clean_text(match.group(1))
            metrics.append(f"{label}: {amount}")
            number_match = re.search(NUMBER_TOKEN, amount)
            if number_match and not re.search(AMOUNT_UNIT_TOKEN, amount, re.IGNORECASE) and "." not in number_match.group(0):
                whole_number_values.append(int(number_match.group(0).replace(",", "")))
    if len(whole_number_values) == 2:
        metrics.append(f"Total demand and penalty: INR {_format_indian_integer(str(sum(whole_number_values)))}")
    return metrics


def _extract_rating_metrics(text: str) -> list[str]:
    metrics: list[str] = []
    unit_match = re.search(
        r"(?:rated amount|amount(?:\s*/\s*size of issue)?).{0,60}"
        r"\(\s*(?:(?:\u20b9|rs\.?|inr)\s*)?(?:in\s+)?([a-z]+)\s*\)",
        text,
        re.IGNORECASE,
    )
    unit = clean_text(unit_match.group(1)).lower() if unit_match else ""
    instrument_pattern = (
        r"(?:Proposed\s+)?Commercial Paper|Long Term Bank Facilities?|Short Term Bank Facilities?|"
        r"Bank Facilities?|Non[- ]Convertible Debentures?|Term Loans?|Bonds?"
    )
    rating_pattern = r"(?:IND|CARE|CRISIL|ICRA|ACUITE|BWR|INFOMERICS)\s+[A-Z]{1,6}[A-Z0-9+()-]*(?:/[A-Za-z]+)?"
    row_pattern = re.compile(
        rf"\b({instrument_pattern})\b(?:\s*\([^)]{{0,80}}\)\s*\*?)?\s+"
        rf"(?:{CURRENCY_TOKEN}\s*)?({NUMBER_TOKEN})\s+({rating_pattern})",
        re.IGNORECASE,
    )
    for match in row_pattern.finditer(text):
        instrument = clean_text(match.group(1))
        value = clean_text(match.group(2))
        rating = clean_text(match.group(3))
        amount_text = f"INR {value}"
        if unit:
            amount_text = f"{amount_text} {unit}"
            if unit in {"million", "millions", "mn"}:
                numeric_value = float(value.replace(",", ""))
                crore_value = numeric_value / 10
                if crore_value.is_integer():
                    crore_text = _format_indian_integer(str(int(crore_value)))
                else:
                    crore_text = f"{crore_value:.2f}".rstrip("0").rstrip(".")
                amount_text = f"{amount_text} (INR {crore_text} crore)"
        metrics.extend([
            f"{instrument} rated amount: {amount_text}",
            f"{instrument} rating: {rating}",
        ])
    return _unique_values(metrics)


def _extract_schedule_metrics(text: str) -> list[str]:
    metrics: list[str] = []
    schedule_match = re.search(
        r"\b((?:[0-9]{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+,?\s+20[0-9]{2}|"
        r"[A-Za-z]+\s+[0-9]{1,2},\s+20[0-9]{2})\s+at\s+"
        r"[0-9]{1,2}:[0-9]{2}\s*(?:a\.?m\.?|p\.?m\.?)?(?:\s*\(?IST\)?)?)",
        text,
        re.IGNORECASE,
    )
    if schedule_match:
        metrics.append(f"Scheduled: {clean_text(schedule_match.group(1))}")
    date_range_match = re.search(
        r"\b([0-9]{1,2}(?:st|nd|rd|th)?\s+and\s+[0-9]{1,2}(?:st|nd|rd|th)?\s+"
        r"[A-Za-z]+,?\s+20[0-9]{2})\b",
        text,
        re.IGNORECASE,
    )
    if date_range_match:
        metrics.append(f"Scheduled: {clean_text(date_range_match.group(1))}")
    quarter_match = re.search(r"\bQ[1-4]FY[0-9]{2,4}\b", text, re.IGNORECASE)
    if quarter_match:
        metrics.append(f"Reporting period: {clean_text(quarter_match.group(0)).upper()}")
    period_match = re.search(r"\b(?:first|second|third|fourth) quarter ended\s+[0-9]{1,2}\s+[A-Za-z]+\s+20[0-9]{2}", text, re.IGNORECASE)
    if period_match:
        metrics.append(f"Results period: {clean_text(period_match.group(0))}")
    return _unique_values(metrics)


def _extract_management_change_metrics(text: str) -> list[str]:
    name_pattern = r"(?:Mr|Ms|Mrs|Smt|Shri)\.?\s+[A-Z][A-Za-z.'-]+(?:\s+[A-Z][A-Za-z.'-]+){1,4}"
    metrics: list[str] = []

    def person_name(value: str) -> str:
        value = re.split(
            r"\b(?:has|have|joined|in place of|with effect|as|who|vide)\b",
            clean_text(value),
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0]
        return value.strip(" ,.;:-")

    for action in ["Appointment", "Resignation", "Cessation"]:
        for match in re.finditer(
            rf"\b{action}\s+of\s+({name_pattern})(?=\s*\(|\s+as\b|[,.;-])",
            text,
            re.IGNORECASE,
        ):
            metrics.append(f"{action}: {person_name(match.group(1))}")
    for match in re.finditer(
        rf"\b({name_pattern}).{{0,100}}?\b(?:has joined|joined)\b.{{0,100}}?\bas\b",
        text,
        re.IGNORECASE,
    ):
        metrics.append(f"Appointment: {person_name(match.group(1))}")
    for match in re.finditer(rf"\binduction of\s+({name_pattern})", text, re.IGNORECASE):
        metrics.append(f"Appointment: {person_name(match.group(1))}")
    for match in re.finditer(rf"\bin place of\s+({name_pattern})", text, re.IGNORECASE):
        metrics.append(f"Replaced committee member: {person_name(match.group(1))}")
    for match in re.finditer(
        rf"\b({name_pattern})(?:\s*\(DIN[^)]*\))?.{{0,180}}?(?:completed (?:his|her) tenure|ceased to hold the office)",
        text,
        re.IGNORECASE,
    ):
        metrics.append(f"Cessation: {person_name(match.group(1))}")
    return _unique_values(metrics)


def _extract_agm_dividend_metrics(text: str) -> list[str]:
    metrics: list[str] = []
    date_token = (
        r"(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)?\s*,?\s*"
        r"(?:[0-9]{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+,?\s+20[0-9]{2}|"
        r"[A-Za-z]+\s+[0-9]{1,2},\s+20[0-9]{2})"
    )
    agm_match = re.search(
        rf"\bAGM\b.{{0,220}}?({date_token})(?:\s*(?:at|\()\s*([0-9]{{1,2}}:[0-9]{{2}}\s*(?:a\.?m\.?|p\.?m\.?)?(?:\s*\(?IST\)?)?))?",
        text,
        re.IGNORECASE,
    )
    if agm_match:
        value = clean_text(agm_match.group(1))
        if agm_match.group(2):
            value = f"{value} at {clean_text(agm_match.group(2))}"
        metrics.append(f"AGM: {value}")
    record_match = re.search(
        rf"\brecord date\b.{{0,120}}?(?:is|fixed for|:)\s*({date_token})",
        text,
        re.IGNORECASE,
    )
    if record_match:
        metrics.append(f"Record date: {clean_text(record_match.group(1))}")
    dividend_match = re.search(
        rf"\b(?:final\s+)?dividend\s+of\s+({CURRENCY_TOKEN}\s*{NUMBER_TOKEN})\s*(?:/-)?\s*per\s+(?:equity\s+)?share",
        text,
        re.IGNORECASE,
    )
    if dividend_match:
        metrics.append(f"Dividend per share: {clean_text(dividend_match.group(1))}")
    return _unique_values(metrics)


def _extract_project_registration_metrics(text: str) -> list[str]:
    metrics: list[str] = []
    for match in re.finditer(
        r"(?:RERA\s+)?registration\s+no\.?\s*[:.-]?\s*([A-Z0-9][A-Z0-9/_-]{8,})",
        text,
        re.IGNORECASE,
    ):
        metrics.append(f"Registration number: {clean_text(match.group(1))}")
    project_match = re.search(r"(?:project,?\s+['\"]|project\s+named\s+)([A-Z][A-Za-z0-9& .'-]{2,80})", text)
    if project_match:
        metrics.append(f"Project: {clean_text(project_match.group(1)).strip(chr(39) + chr(34))}")
    return _unique_values(metrics)


def _extract_operating_volume_metrics(text: str) -> list[str]:
    start_match = re.search(r"\bthroughput\b", text, re.IGNORECASE)
    table_text = text[start_match.start(): start_match.start() + 1400] if start_match else text
    label_matches = list(re.finditer(r"\b(?:EXIM|DOM|Total)\b", table_text, re.IGNORECASE))
    metrics: list[str] = []
    for index, label_match in enumerate(label_matches):
        end = label_matches[index + 1].start() if index + 1 < len(label_matches) else min(len(table_text), label_match.start() + 260)
        segment = table_text[label_match.start():end]
        counts = [
            _format_indian_integer(match.group(1))
            for match in INDIAN_GROUPED_NUMBER_PATTERN.finditer(segment)
        ]
        variation_match = re.search(rf"([+-]?[0-9]+(?:\.[0-9]+)?)\s*{PERCENT_TOKEN}", segment, re.IGNORECASE)
        if len(counts) >= 2 and variation_match:
            label = clean_text(label_match.group(0)).upper()
            variation = clean_text(variation_match.group(1))
            metrics.append(f"{label} throughput: {counts[0]} TEUs vs {counts[1]} TEUs ({variation}% variation)")
    return _unique_values(metrics)


def _first_pattern_match(text: str, patterns: list[str]):
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            return match
    return None


def _event_evidence(text: str, patterns: list[str], chars: int = 650) -> str:
    match = _first_pattern_match(text, patterns)
    return context_window(text, match.start(), match.end(), chars) if match else ""


def _unique_values(values: list[str]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = clean_text(value)
        key = cleaned.lower()
        if cleaned and key not in seen:
            seen.add(key)
            output.append(cleaned)
    return output


def _extract_target_counterparty(text: str) -> str:
    patterns = [
        r"name of (?:the )?target entity.{0,180}?[;:]\s*([A-Z][A-Za-z0-9&().,' -]{2,100}?(?:Private Limited|Limited|Ltd\.?))",
        r"acquisition of\s+[0-9.]+%\s+of (?:the )?equity share capital of\s+([A-Z][A-Za-z0-9&().,' -]{2,100}?(?:Private Limited|Limited|Ltd\.?))",
        r"(?:acquisition of|acquired)\s+(?:a\s+stake\s+in\s+)?([A-Z][A-Za-z0-9&().,' -]{2,100}?(?:Private Limited|Limited|Ltd\.?))",
        r"(?:agreement|memorandum of understanding)\s+with\s+([A-Z][A-Za-z0-9&().,' -]{2,100}?(?:Private Limited|Limited|Ltd\.?))(?=,|\s+to\b|\s+for\b|[.;]|$)",
        r"(?:agreement|memorandum of understanding)\s+with\s+([A-Z][A-Za-z0-9&().,' -]{2,100})",
        r"by the name(?: of)?\s+([A-Z][A-Za-z0-9&().,' -]{2,100}?(?:Private Limited|Limited|Ltd\.?))",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            value = clean_text(match.group(1))
            value = re.split(
                r",?\s+(?:to\s+(?:establish|develop|set\s+up|operate|provide|create)\b|for\s+(?:the\s+)?(?:establishment|development|setting\s+up)\b)",
                value,
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0]
            value = re.split(r"\s{2,}|\b(?:details|whether|turnover|percentage|industry)\b", value, maxsplit=1, flags=re.IGNORECASE)[0]
            return value.strip(" .,:;-")[:140]
    return ""


def _extract_event_metrics(text: str) -> list[str]:
    metrics: list[str] = []
    metric_text = re.sub(r"(?<=[0-9])[lL](?=[0-9,])", "1", text)
    for match in re.finditer(
        r"(?<![0-9])([0-9]{1,3}(?:\s*,\s*[0-9]{2})*\s*,\s*[0-9]{3}|[0-9]+)\s+"
        r"((?:equity\s+)?shares)\b",
        metric_text,
        re.IGNORECASE,
    ):
        number = re.sub(r"\s*,\s*", ",", match.group(1))
        metrics.append(f"{number} {clean_text(match.group(2))}")
    patterns = [
        rf"\b[0-9]+(?:\.[0-9]+)?\s*{PERCENT_TOKEN}",
        r"\b[0-9]+(?:\.[0-9]+)?\s*(?:acres?|acre)\b",
        r"\b[0-9]+(?:\.[0-9]+)?[- ]year\b",
        r"\b[0-9][0-9,.]*\s*(?:square\s+met(?:er|re)s?|sq\.?\s*m(?:et(?:er|re)s?)?)\b",
        r"\b[0-9][0-9,]*\s+(?:LNG[\s-]*p\s*owered\s+)?commercial vehicles?\b",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, metric_text, re.IGNORECASE):
            metric = clean_text(match.group(0))
            metric = re.sub(PERCENT_TOKEN, "%", metric, flags=re.IGNORECASE)
            metric = re.sub(r"LNG[\s-]*p\s*owered", "LNG-powered", metric, flags=re.IGNORECASE)
            if re.match(r"^[0-9]{1,3}\.[0-9]{3}\s+square", metric, re.IGNORECASE):
                metric = metric.replace(".", ",", 1)
            metrics.append(metric)
    return _unique_values(metrics)[:12]


def classify_pdf_result(pdf_path: Path, candidate: dict[str, Any], args: argparse.Namespace) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    """Classify a disclosure using completed-action language, not loose keywords."""
    text, pages, parse_status = extract_pdf_text(pdf_path)
    scan = clean_text(text)
    row_context = clean_text(candidate.get("context"))
    subject = extract_subject(text, fallback=row_context or pdf_path.stem)
    descriptor = clean_text(" ".join([pdf_path.name, subject, row_context, scan[:1800]]))
    descriptor_lower = descriptor.lower()
    direct_descriptor = clean_text(" ".join([pdf_path.name, subject, row_context]))
    direct_descriptor_lower = direct_descriptor.lower()
    direct_intent = clean_text(re.sub(r"[_-]+", " ", direct_descriptor))

    document_type = "Other Disclosure"
    infusion_type = "None"
    capital = "No"
    debt = "No"
    esop_only = "No"
    esop_related = "No"
    cash_disclosed = "Not applicable"
    material_event = "No"
    material_event_type = ""
    event_status = "No event"
    target_counterparty = ""
    confidence = "High"
    note = "No completed capital infusion or debt raising detected in this disclosure."
    reasoning = "No completed-action financing language was found."
    evidence = ""
    amount_source = ""
    key_numbers: list[str] = []
    financial_relevance = "No"
    financial_relevance_level = "None"
    financial_category = "No material financial event detected"
    financial_note = "PDF parsed and listed; no specific financial or material event was detected by the rules."

    intent_scan = clean_text(" ".join([direct_intent, scan[:5000]]))
    acquisition_scan = scan[:9000]
    tax_scan = clean_text(" ".join([direct_intent, scan[:4500]]))
    is_notice = bool(re.search(
        r"newspaper|advertisement|publication|special window|re-lodgement|postal ballot|remote e-voting",
        direct_intent,
        re.IGNORECASE,
    ))
    is_financial_results_clarification = bool(re.search(
        r"(?:reply to\s+)?clarification.{0,180}financial results|financial results.{0,180}clarification",
        direct_intent,
        re.IGNORECASE,
    ))
    is_brsr = "brsr" in direct_descriptor_lower or "business responsibility and sustainability report" in descriptor_lower
    is_annual_report = bool(
        not is_brsr
        and not is_notice
        and re.search(r"annual report", direct_intent, re.IGNORECASE)
        and not re.search(r"weblink|letter to (?:members|shareholders)|communication to shareholders", direct_intent, re.IGNORECASE)
        and not is_financial_results_clarification
    )
    is_earnings_call = bool(re.search(
        r"earnings (?:conference )?call|analysts?/institutional investor|investor (?:meet|meeting|call)|con\.\s*call updates",
        descriptor,
        re.IGNORECASE,
    ))
    is_shareholder_annual_report_communication = bool(re.search(
        r"letter to (?:members|shareholders)|weblink.{0,240}(?:annual report|annual general meeting)|regulation\s+36\s*\(?1\)?",
        descriptor,
        re.IGNORECASE,
    )) and bool(re.search(r"annual report|annual general meeting|\bAGM\b", descriptor, re.IGNORECASE))
    is_agm_dividend_notice = bool(
        not is_annual_report
        and re.search(r"annual general meeting|\bAGM\b", direct_intent, re.IGNORECASE)
        and re.search(r"record date|final dividend|declaration of dividend", intent_scan, re.IGNORECASE)
    )
    is_dividend_tds = bool(re.search(
        r"tax deduction at source|withholding tax|\bTDS\b|communication.{0,100}dividend|dividend.{0,100}communication",
        clean_text(" ".join([direct_intent, scan[:2800]])),
        re.IGNORECASE,
    ))
    is_operating_volume_update = bool(re.search(
        r"physical volumes?|throughput\s*\(?(?:in\s+)?TEUs?\)?|volumes? handled",
        descriptor,
        re.IGNORECASE,
    ))
    takeover_scan = clean_text(" ".join([direct_intent, scan[:3200]]))
    is_takeover_disclosure = bool(
        re.search(
            r"regulation\s+31\s*\(?4\)?|non[- ]encumbrance|exceeded\s+5%|takeover regulations|"
            r"substantial acquisition of shares|\bSAST\b",
            takeover_scan,
            re.IGNORECASE,
        )
        or (
            re.search(r"regulation\s+29\s*\(?[12]?\)?", takeover_scan, re.IGNORECASE)
            and re.search(r"substantial acquisition|takeover|\bSAST\b", takeover_scan, re.IGNORECASE)
        )
    )
    is_governance_change = bool(
        re.search(r"appointment|resignation|cessation|change in management", direct_intent, re.IGNORECASE)
        or re.search(
            r"reconstitution.{0,120}(?:committee|NRC)|induction of.{0,120}in place of|"
            r"(?:Mr|Ms|Mrs|Smt|Shri)\.?\s+[A-Z][A-Za-z.'-]+(?:\s+[A-Z][A-Za-z.'-]+){1,4}."
            r"{0,100}(?:has joined|joined).{0,100}\bas\b|"
            r"appointment of\s+(?:Mr|Ms|Mrs|Smt|Shri)\.?\s+[A-Z]",
            scan[:3500],
            re.IGNORECASE,
        )
    )
    is_project_regulatory_update = bool(
        re.search(r"\bproject\b|\bRERA\b", intent_scan, re.IGNORECASE)
        and re.search(
            r"transfer of rights|\bTOR\b.{0,120}registration|RERA.{0,160}(?:registration|approval)|registration certificate",
            intent_scan,
            re.IGNORECASE,
        )
    )

    esop_exercise_patterns = [
        r"allot(?:ted|ment of)\s+[0-9,]+\s+equity shares.{0,420}(?:exercise|employee stock|esop|option)",
        r"(?:upon|pursuant to)\s+(?:the\s+)?exercise.{0,260}(?:vested\s+)?(?:option|esop)",
        r"allot(?:ted|ment).{0,220}equity shares.{0,320}(?:esop|stock option|exercise of option)",
    ]
    esop_grant_patterns = [
        r"grant of\s+[0-9,]+\s+(?:employee\s+)?stock options?",
        r"[0-9,]+\s+(?:employee\s+)?stock options?\s+(?:were\s+)?granted",
        r"approval.{0,120}grant.{0,120}stock options?",
    ]
    direct_esop_grant = bool(re.search(r"grant.{0,80}(?:employee\s+)?stock options?|stock options?.{0,80}grant", direct_intent, re.IGNORECASE))
    subsidiary_subscription_patterns = [
        r"(?:has|have)\s+subscribed.{0,500}(?:equity shares|share capital)",
        r"(?:equity shares|share capital).{0,500}(?:has|have)\s+subscribed",
        r"by way of subscription.{0,500}(?:equity shares|memorandum of association)",
        r"cost of (?:acquisition|subscription).{0,500}(?:cash consideration|equity shares|inr|rs\.?|\u20b9)",
    ]
    proposed_subsidiary_patterns = [
        r"approved.{0,240}incorporat(?:e|ion).{0,300}(?:subsidiar|wholly owned)",
        r"incorporat(?:e|ion).{0,260}(?:subsidiar|wholly owned).{0,320}shall subscribe",
        r"shall subscribe.{0,320}(?:subsidiar|wholly owned|equity shares)",
    ]
    spa_patterns = [
        r"entered into.{0,180}(?:share purchase agreement|share subscription agreement)",
        r"executed.{0,180}(?:share purchase agreement|share subscription agreement)",
    ]
    acquisition_patterns = [
        r"\bhas acquired\b.{0,420}(?:stake|shares|business|company|land|parcel)",
        r"\bacquires?\b.{0,260}(?:stake|shares|business|company|land|parcel)",
        r"acquisition (?:has been|was) completed",
        r"acquisition of.{0,260}(?:acre|land parcel|equity stake|equity shares)",
    ]
    rating_patterns = [
        r"\bcredit rating(?:s)?\b",
        r"\brating action\b",
        r"\brating(?:s)?\s+(?:assigned|reaffirmed|upgraded|downgraded|withdrawn)\b",
        r"(?:commercial paper|bank facilities?).{0,220}\b(?:rated|rating)\b",
    ]
    debt_instruments = r"(?:commercial paper|non[- ]convertible debentures?|ncds?|debt securities?|term loans?|working capital loans?|credit facilit(?:y|ies)|bank facilit(?:y|ies)|borrowings?|bonds?)"
    debt_actions = r"(?:issued|allotted|raised|availed|borrowed|drawn(?:\s+down)?|disbursed)"
    actual_debt_patterns = [
        rf"\b{debt_actions}\b.{{0,180}}\b{debt_instruments}\b",
        rf"\b{debt_instruments}\b.{{0,180}}\b(?:has|have|was|were)?\s*{debt_actions}\b",
    ]
    proposed_debt_patterns = [
        rf"\b(?:proposed|proposal|approved|approval|considered)\b.{{0,220}}\b{debt_instruments}\b",
        rf"\b{debt_instruments}\b.{{0,180}}\b(?:proposed|to be issued|subject to approval)\b",
    ]
    actual_equity_patterns = [
        r"\b(?:has|have|were|was)?\s*(?:issued|allotted)\b.{0,180}\b(?:equity shares|preference shares|rights shares)\b",
        r"\b(?:equity shares|preference shares|rights shares)\b.{0,180}\b(?:were|have been|has been)\s+(?:issued|allotted)\b",
    ]
    proposed_equity_patterns = [
        r"\b(?:approved|proposed|considered)\b.{0,220}\b(?:rights issue|preferential issue|preferential allotment|issue of equity shares)\b",
        r"\b(?:rights issue|preferential issue|issue of equity shares)\b.{0,180}\b(?:proposed|subject to approval)\b",
    ]
    tax_order_patterns = [
        r"(?:gst|tax)\s+(?:demand|order|penalty)",
        r"(?:demand|penalty).{0,100}(?:gst|tax)\b",
        r"\bpenalty (?:has been |was )?(?:imposed|levied)\b",
        r"\border.{0,100}(?:passed|received).{0,120}(?:gst|tax)\b",
    ]

    if parse_status != "Parsed" and is_takeover_disclosure:
        document_type = "Shareholding / Takeover Disclosure"
        event_status = "Compliance filing"
        confidence = "Medium"
        evidence = row_context
        note = "Takeover-code compliance filing identified from the NSE announcement row; PDF content still requires manual review."
        reasoning = "The image-only PDF has no selectable text, but the NSE row explicitly identifies Regulation 31(4)/takeover compliance."
        financial_relevance = "Yes"
        financial_relevance_level = "Direct securities-related (announcement metadata)"
        financial_category = "Shareholding / takeover disclosure"
        financial_note = "Securities ownership or takeover-code filing; no company capital or debt infusion is inferred."
        material_event_type = "Shareholding compliance"
    elif parse_status != "Parsed":
        document_type = "Parse Issue / Image-only PDF"
        infusion_type = "Review needed"
        event_status = "Review needed"
        confidence = "Low"
        note = parse_status
        reasoning = "The PDF is retained in the workbook, but it has no extractable text and requires manual review."
        financial_relevance = "Unknown"
        financial_relevance_level = "Review needed"
        financial_category = "Parse issue / image-only PDF"
        financial_note = "Financial relevance cannot be determined automatically because selectable text was unavailable."
    elif is_brsr:
        document_type = "BRSR / ESG Report"
        event_status = "Report filed"
        financial_relevance = "Yes"
        financial_relevance_level = "Indirect reporting"
        financial_category = "BRSR / ESG reporting"
        financial_note = "The report is relevant for operating and ESG review, but incidental financing or ESOP references are not current events."
        note = "BRSR filing; no completed capital or debt event is counted from historical narrative."
        reasoning = "Long-form reports contain historical references, so only a direct event filing can create an infusion flag."
    elif is_financial_results_clarification:
        document_type = "Financial Results Clarification"
        event_status = "Clarification filed"
        evidence = _event_evidence(
            scan,
            [r"clarification.{0,180}financial results", r"financial results.{0,180}clarification", r"consolidated PAT|EPS|XBRL"],
            chars=1000,
        ) or scan[:1000]
        financial_relevance = "Yes"
        financial_relevance_level = "Direct reporting clarification"
        financial_category = "Financial results clarification"
        financial_note = "Exchange clarification concerning reported financial results or XBRL data."
        note = "Financial-results clarification filed; no capital/debt infusion."
        reasoning = "The NSE subject explicitly identifies a clarification to published financial results."
    elif is_annual_report:
        document_type = "Annual Report"
        event_status = "Report filed"
        financial_relevance = "Yes"
        financial_relevance_level = "Direct reporting"
        financial_category = "Annual financial report"
        financial_note = "Annual financial report included for review; historical financing and ESOP references are not treated as new events."
        note = "Annual report filing; no completed current-period infusion event is inferred from incidental text."
        reasoning = "Long-form annual reports are kept as reports instead of being misclassified from isolated keywords."
    elif is_earnings_call:
        document_type = "Earnings Call / Investor Meeting"
        event_status = "Scheduled"
        evidence = _event_evidence(
            scan,
            [r"earnings (?:conference )?call", r"analysts?/institutional investor", r"investor (?:meet|meeting|call)"],
            chars=1200,
        ) or scan[:1200]
        key_numbers.extend(_extract_schedule_metrics(scan))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct results-related communication"
        financial_category = "Earnings call / analyst interaction"
        financial_note = "Scheduled investor interaction concerning financial results; no financing event."
        note = "Earnings call or analyst/investor meeting scheduled."
        reasoning = "The filing schedules a results-related investor call and is financially relevant even though it is not a capital/debt event."
    elif is_notice:
        document_type = "Shareholder / Newspaper Notice"
        event_status = "Notice"
        evidence = _event_evidence(
            scan,
            [r"newspaper publication|newspaper advertisement|postal ballot|remote e-voting|special window|re-lodgement"],
            chars=900,
        ) or scan[:900]
        key_numbers.extend(_extract_agm_dividend_metrics(scan[:5000]))
        financial_relevance = "Yes"
        financial_relevance_level = "Shareholder-related"
        financial_category = "Shareholder voting / publication"
        financial_note = "Governance/shareholder notice; no capital/debt infusion."
        note = "Shareholder or newspaper notice."
        reasoning = "The direct NSE subject identifies a notice/publication, so unrelated article text cannot reclassify it."
    elif is_shareholder_annual_report_communication:
        document_type = "Shareholder / Annual Report Communication"
        event_status = "Notice"
        evidence = _event_evidence(
            scan,
            [r"letter to (?:members|shareholders)", r"weblink.{0,240}(?:annual report|annual general meeting)", r"regulation\s+36\s*\(?1\)?"],
            chars=1000,
        ) or scan[:1000]
        financial_relevance = "Yes"
        financial_relevance_level = "Shareholder / annual reporting"
        financial_category = "Annual report / AGM shareholder communication"
        financial_note = "Shareholder communication providing access to the AGM notice and annual financial report."
        note = "Annual-report/AGM communication to shareholders."
        reasoning = "The document distributes or links the annual report and AGM notice rather than announcing a financing event."
    elif is_agm_dividend_notice:
        document_type = "AGM / Dividend Record Date Notice"
        event_status = "Notice"
        evidence = _event_evidence(scan, [r"annual general meeting|\bAGM\b|record date|final dividend"], chars=1100) or scan[:1100]
        key_numbers.extend(_extract_agm_dividend_metrics(scan))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct shareholder financial"
        financial_category = "AGM / dividend / record date"
        financial_note = "AGM and dividend record-date communication; no capital/debt infusion."
        note = "AGM/dividend record-date notice."
        reasoning = "A meeting and record-date notice is not an annual-report filing or financing event."
    elif is_operating_volume_update:
        document_type = "Operating Metrics / Physical Volumes"
        event_status = "Reported/Provisional"
        material_event = "Yes"
        material_event_type = "Operating performance update"
        evidence = _event_evidence(scan, [r"physical volumes?", r"throughput", r"volumes? handled"], chars=1200)
        key_numbers.extend(_extract_operating_volume_metrics(scan))
        if not key_numbers:
            key_numbers.extend(_extract_event_metrics(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct operating performance"
        financial_category = "Operating volumes / throughput"
        financial_note = "Current operating volumes and year-on-year changes are direct business-performance indicators."
        note = "Operating-volume update reported."
        reasoning = "The disclosure reports current throughput KPIs, so it is financially relevant even without a capital/debt event."
    elif is_takeover_disclosure:
        document_type = "Shareholding / Takeover Disclosure"
        is_non_encumbrance = bool(re.search(r"regulation\s+31\s*\(?4\)?|non[- ]encumbrance", intent_scan, re.IGNORECASE))
        is_actual_change = bool(
            not is_non_encumbrance
            and re.search(r"regulation\s+29|exceeded\s+5%|net acquisition|shareholding.{0,160}(?:increased|decreased)", intent_scan, re.IGNORECASE)
        )
        event_status = "Actual" if is_actual_change else "Compliance filing"
        material_event = "Yes" if is_actual_change else "No"
        material_event_type = "Significant shareholding change" if is_actual_change else "Shareholding compliance"
        evidence = _event_evidence(
            scan,
            [r"regulation\s+31\s*\(?4\)?|non[- ]encumbrance|regulation\s+29|exceeded\s+5%|net acquisition"],
            chars=1300,
        ) or scan[:1300]
        key_numbers.extend(extract_amounts(evidence))
        key_numbers.extend(_extract_event_metrics(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct securities-related"
        financial_category = "Shareholding / takeover disclosure"
        financial_note = "Securities ownership or takeover-code filing; not automatically a company capital infusion."
        note = "Shareholding/takeover disclosure; financing flags remain off."
        reasoning = "Regulation 31(4) non-encumbrance filings are annual compliance, not acquisitions."
    elif is_dividend_tds:
        document_type = "Dividend / TDS Communication"
        event_status = "Notice"
        evidence = _event_evidence(scan, [r"tax deduction at source|\bTDS\b|withholding tax|\bdividend\b"], chars=1000) or scan[:1000]
        key_numbers.extend(_extract_agm_dividend_metrics(scan))
        key_numbers.extend(extract_amounts(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct shareholder financial"
        financial_category = "Dividend / withholding tax"
        financial_note = "Dividend or TDS communication is financially relevant but is not a tax order or financing event."
        note = "Dividend/TDS shareholder communication."
        reasoning = "TDS guidance for a dividend is a shareholder communication, not a tax demand or penalty."
    elif is_governance_change:
        document_type = "Management Change"
        event_status = "Actual"
        material_event = "Yes"
        material_event_type = "Management / governance change"
        evidence = _event_evidence(
            scan,
            [r"appointment|resignation|cessation|change in management|reconstitution|induction of|has joined"],
            chars=1000,
        ) or scan[:1000]
        key_numbers.extend(_extract_management_change_metrics(scan))
        financial_relevance = "Yes"
        financial_relevance_level = "Indirect governance"
        financial_category = "Management / governance change"
        financial_note = "Governance or senior-management event; no capital/debt infusion."
        note = "Management or board-committee change disclosure."
        reasoning = "Appointments, cessations, and committee reconstitutions are governance events."
    elif is_project_regulatory_update:
        document_type = "Project / RERA Update"
        event_status = "Actual"
        material_event = "Yes"
        material_event_type = "Project / regulatory milestone"
        evidence = _event_evidence(
            scan,
            [r"transfer of rights|\bTOR\b.{0,120}registration|RERA.{0,160}(?:registration|approval)|registration certificate"],
            chars=1100,
        ) or scan[:1100]
        key_numbers.extend(_extract_project_registration_metrics(scan))
        financial_relevance = "Yes"
        financial_relevance_level = "Indirect commercial"
        financial_category = "Project / RERA update"
        financial_note = "Project registration or regulatory milestone relevant to operating activity."
        note = "Project/RERA regulatory update."
        reasoning = "The filing reports a current project registration milestone rather than a financing event."
    elif _first_pattern_match(scan, esop_exercise_patterns) and not direct_esop_grant:
        document_type = "ESOP Exercise / Share Allotment"
        infusion_type = "ESOP share allotment / paid-up capital increase"
        capital = "Yes"
        esop_related = "Yes"
        cash_disclosed = "Not disclosed"
        material_event = "Yes"
        material_event_type = "Capital increase through ESOP exercise"
        event_status = "Actual"
        evidence = _event_evidence(scan, esop_exercise_patterns, chars=1100)
        amount_source = scan
        allotted_match = re.search(
            r"allot(?:ted|ment of)(?:\s+(?:a\s+)?total\s+of)?\s+([0-9][0-9,]*)(?:\s+\([^)]*\))?\s+equity shares",
            scan,
            re.IGNORECASE,
        )
        if allotted_match:
            key_numbers.append(f"Equity shares allotted: {allotted_match.group(1)}")
        capital_match = re.search(
            rf"paid\s*-?\s*up(?: equity)? share capital.{{0,180}}?increased from\s*{CURRENCY_TOKEN}\s*([0-9,]+).{{0,100}}?\bto\s*{CURRENCY_TOKEN}\s*([0-9,]+)",
            scan,
            re.IGNORECASE,
        )
        if capital_match:
            before = int(capital_match.group(1).replace(",", ""))
            after = int(capital_match.group(2).replace(",", ""))
            key_numbers.extend([
                f"Paid-up capital before: INR {_format_indian_integer(str(before))}",
                f"Paid-up capital after: INR {_format_indian_integer(str(after))}",
                f"Paid-up capital increase: INR {_format_indian_integer(str(after - before))}",
            ])
        for option_count, exercise_price in re.findall(
            r"\b([0-9][0-9,]*)\s+Options?\s*@\s*((?:Re\s*\.?|Rs\s*\.?|\u20b9|INR)\s*[0-9][0-9,]*(?:\.[0-9]+)?)",
            scan,
            re.IGNORECASE,
        )[:6]:
            key_numbers.append(f"Exercise price: {clean_text(exercise_price)} for {option_count} options")
        money_match = re.search(
            rf"money reali[sz]ed(?: by exercise of options)?.{{0,120}}?({CURRENCY_TOKEN}\s*{NUMBER_TOKEN})",
            scan,
            re.IGNORECASE,
        )
        if money_match:
            cash_disclosed = "Yes"
            key_numbers.append(f"Money realized on exercise: {clean_text(money_match.group(1))}")
        financial_relevance = "Yes"
        financial_relevance_level = "Direct securities / capital"
        financial_category = "ESOP exercise and paid-up capital increase"
        financial_note = "Shares were actually allotted following option exercise; paid-up capital increased. Cash proceeds are only marked disclosed when stated."
        note = "Actual equity-share allotment following ESOP exercise."
        reasoning = "Completed allotment/exercise language confirms an actual capital increase, unlike a grant of options."
    elif _first_pattern_match(scan, esop_grant_patterns):
        document_type = "ESOP / Stock Option Grant"
        infusion_type = "ESOP grant only"
        esop_only = "Yes"
        esop_related = "Yes"
        event_status = "Granted"
        material_event = "Yes"
        material_event_type = "ESOP grant"
        evidence = _event_evidence(scan, esop_grant_patterns, chars=700)
        option_counts = re.findall(r"\b([0-9][0-9,]*)\s+(?:employee\s+)?stock options?\b", evidence, re.IGNORECASE)
        key_numbers.extend([f"Stock options granted: {value}" for value in _unique_values(option_counts)[:3]])
        key_numbers.extend(_extract_event_metrics(evidence))
        key_numbers.extend(extract_amounts(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct securities-related"
        financial_category = "ESOP / potential dilution"
        financial_note = "Options were granted, but no capital is counted until exercise and share allotment."
        note = "Stock-option grant only; no actual capital infusion."
        reasoning = "A grant creates potential dilution but does not itself issue shares or raise cash."
    elif _first_pattern_match(scan, subsidiary_subscription_patterns) and re.search(r"subsidiar|wholly owned|target entity|equity shares", scan, re.IGNORECASE):
        document_type = "Subsidiary Investment / Share Subscription"
        infusion_type = "Capital subscription / subsidiary investment"
        capital = "Yes"
        material_event = "Yes"
        material_event_type = "Subsidiary investment / capital subscription"
        event_status = "Actual"
        target_counterparty = _extract_target_counterparty(scan)
        evidence = _event_evidence(scan, subsidiary_subscription_patterns, chars=950)
        cash_disclosed = "Yes" if re.search(r"cash consideration|(?:\u20b9|rs\.?|inr)\s*[0-9]", evidence, re.IGNORECASE) else "Not disclosed"
        share_counts = re.findall(r"\b([0-9][0-9,]*)\s+equity shares\b", evidence, re.IGNORECASE)
        key_numbers.extend([f"Equity shares subscribed/acquired: {value}" for value in _unique_values(share_counts)[:3]])
        key_numbers.extend(_extract_event_metrics(evidence))
        key_numbers.extend(extract_amounts(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct investment / capital"
        financial_category = "Subsidiary investment / equity subscription"
        financial_note = "Completed subscription language confirms an actual equity investment in the subsidiary/target."
        note = "Actual share subscription or subsidiary capital investment detected."
        reasoning = "The disclosure states that shares were subscribed/acquired and identifies consideration or a completed subscription."
    elif _first_pattern_match(scan, proposed_subsidiary_patterns):
        document_type = "Proposed Subsidiary Incorporation"
        infusion_type = "Proposed capital subscription"
        material_event = "Yes"
        material_event_type = "Subsidiary incorporation / proposed subscription"
        event_status = "Approved/Proposed"
        target_counterparty = _extract_target_counterparty(scan)
        evidence = _event_evidence(scan, proposed_subsidiary_patterns, chars=850)
        key_numbers.extend(extract_amounts(evidence))
        key_numbers.extend(_extract_event_metrics(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct proposed investment"
        financial_category = "Proposed subsidiary / future subscription"
        financial_note = "The board approved incorporation or a future subscription; it is not counted as completed capital infusion."
        note = "Approved/proposed subsidiary investment; not yet an actual infusion."
        reasoning = "Future-tense or approval language is separated from completed subscription language."
    elif _first_pattern_match(acquisition_scan, spa_patterns):
        document_type = "Acquisition / Share Purchase Agreement"
        material_event = "Yes"
        material_event_type = "Acquisition agreement"
        event_status = "Contracted/Proposed"
        target_counterparty = _extract_target_counterparty(scan)
        evidence = _event_evidence(acquisition_scan, spa_patterns, chars=900)
        key_numbers.extend(extract_amounts(evidence))
        key_numbers.extend(_extract_event_metrics(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct strategic transaction"
        financial_category = "M&A / investment agreement"
        financial_note = "A share-purchase/subscription agreement was executed; completion and capital infusion are not assumed unless separately stated."
        note = "Material acquisition agreement detected; no completed capital/debt infusion counted."
        reasoning = "Executing an SPA is material, but it does not by itself prove that consideration was paid or the acquisition completed."
    elif (
        _first_pattern_match(acquisition_scan, acquisition_patterns)
        and not is_takeover_disclosure
        and not is_notice
        and not is_dividend_tds
    ):
        document_type = "Acquisition / Strategic Investment"
        material_event = "Yes"
        material_event_type = "Acquisition / land or equity investment"
        event_status = "Actual"
        target_counterparty = _extract_target_counterparty(scan)
        evidence = _event_evidence(acquisition_scan, acquisition_patterns, chars=950)
        key_numbers.extend(extract_amounts(evidence))
        key_numbers.extend(_extract_event_metrics(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct strategic transaction"
        financial_category = "M&A / land / equity acquisition"
        financial_note = "A completed acquisition or land investment was announced. It is material but is not automatically debt or capital infusion."
        note = "Actual acquisition/investment detected; kept separate from financing flags."
        reasoning = "Completed acquisition language confirms a strategic transaction, not necessarily a fundraising event."
    elif _first_pattern_match(descriptor, rating_patterns) or _first_pattern_match(scan[:12000], rating_patterns):
        document_type = "Credit Rating / Rated Facilities"
        infusion_type = "Rating only"
        event_status = "Rating only"
        evidence = _event_evidence(scan, rating_patterns + [r"commercial paper", r"bank facilities?"], chars=1000)
        amount_source = scan
        rating_metrics = _extract_rating_metrics(scan)
        if rating_metrics:
            key_numbers.extend(rating_metrics)
        else:
            key_numbers.extend(extract_amounts(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct financing-related"
        financial_category = "Credit rating / proposed or existing facilities"
        financial_note = "Rated limits or a proposed instrument are financially relevant, but no new borrowing is counted without issuance/drawdown language."
        note = "Credit-rating disclosure; no actual debt infusion counted."
        reasoning = "A rating or rated limit is not evidence that funds were issued, borrowed, or drawn."
    elif _first_pattern_match(scan, actual_debt_patterns):
        document_type = "Debt Issuance / Borrowing"
        infusion_type = "Debt infusion"
        debt = "Yes"
        material_event = "Yes"
        material_event_type = "Debt raised / facility drawn"
        event_status = "Actual"
        evidence = _event_evidence(scan, actual_debt_patterns, chars=900)
        key_numbers.extend(extract_amounts(evidence))
        cash_disclosed = "Yes" if extract_amounts(evidence) else "Not disclosed"
        financial_relevance = "Yes"
        financial_relevance_level = "Direct financing"
        financial_category = "Actual debt issuance / borrowing"
        financial_note = "Completed issuance, borrowing, or drawdown language confirms an actual debt event."
        note = "Actual debt raising detected."
        reasoning = "The financing instrument appears next to a completed action such as issued, raised, availed, borrowed, or drawn."
    elif _first_pattern_match(scan, proposed_debt_patterns):
        document_type = "Proposed Debt / Financing Approval"
        infusion_type = "Proposed debt"
        material_event = "Yes"
        material_event_type = "Proposed debt raising"
        event_status = "Approved/Proposed"
        evidence = _event_evidence(scan, proposed_debt_patterns, chars=900)
        key_numbers.extend(extract_amounts(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct proposed financing"
        financial_category = "Proposed debt / financing approval"
        financial_note = "A debt issue or facility is proposed/approved but is not counted until issuance or drawdown."
        note = "Proposed debt financing; no actual debt infusion counted."
        reasoning = "Proposal and approval language is not a completed borrowing action."
    elif _first_pattern_match(scan, actual_equity_patterns):
        document_type = "Equity Share Issuance / Allotment"
        infusion_type = "Equity capital increase"
        capital = "Yes"
        material_event = "Yes"
        material_event_type = "Equity issuance / allotment"
        event_status = "Actual"
        evidence = _event_evidence(scan, actual_equity_patterns, chars=900)
        key_numbers.extend(extract_amounts(evidence))
        share_counts = re.findall(r"\b([0-9][0-9,]*)\s+(?:equity|preference|rights) shares\b", evidence, re.IGNORECASE)
        key_numbers.extend([f"Shares issued/allotted: {value}" for value in _unique_values(share_counts)[:3]])
        cash_disclosed = "Yes" if extract_amounts(evidence) else "Not disclosed"
        financial_relevance = "Yes"
        financial_relevance_level = "Direct capital"
        financial_category = "Actual equity issuance / allotment"
        financial_note = "Completed share issuance/allotment language confirms an actual capital event."
        note = "Actual equity issuance or allotment detected."
        reasoning = "The disclosure uses completed issuance/allotment language."
    elif _first_pattern_match(scan, proposed_equity_patterns):
        document_type = "Proposed Equity / Capital Approval"
        infusion_type = "Proposed equity capital"
        material_event = "Yes"
        material_event_type = "Proposed equity issuance"
        event_status = "Approved/Proposed"
        evidence = _event_evidence(scan, proposed_equity_patterns, chars=850)
        key_numbers.extend(extract_amounts(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct proposed capital"
        financial_category = "Proposed equity issuance"
        financial_note = "The issue is proposed or approved but not counted as actual capital until allotment/issuance."
        note = "Proposed equity event; no actual capital infusion counted."
        reasoning = "Proposal/approval language is separated from completed issuance."
    elif _first_pattern_match(tax_scan, tax_order_patterns) and not is_notice and not is_dividend_tds:
        document_type = "Regulatory / Tax Order"
        material_event = "Yes"
        material_event_type = "Tax demand / penalty / order"
        event_status = "Actual"
        evidence = _event_evidence(scan, tax_order_patterns, chars=850)
        amount_source = evidence
        tax_metrics = _extract_tax_metrics(evidence)
        if tax_metrics:
            key_numbers.extend(tax_metrics)
        else:
            key_numbers.extend(extract_amounts(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct financial"
        financial_category = "Tax / GST / penalty"
        financial_note = "An explicit tax/GST demand, order, or imposed penalty was detected. A GST registration number alone cannot trigger this category."
        note = "Tax/GST order or penalty; no capital/debt infusion."
        reasoning = "Explicit order, demand, or imposed-penalty language is required; boilerplate GSTN text is ignored."
    elif is_dividend_tds:
        document_type = "Dividend / TDS Communication"
        event_status = "Notice"
        evidence = _event_evidence(scan, [r"tax deduction at source|\btds\b|\bdividend\b"], chars=700)
        key_numbers.extend(extract_amounts(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct shareholder financial"
        financial_category = "Dividend / withholding tax"
        financial_note = "Dividend or TDS communication is financially relevant but is not a capital/debt infusion."
        note = "Dividend/TDS shareholder communication."
        reasoning = "The disclosure concerns shareholder distribution or withholding tax."
    elif "trading window" in descriptor_lower:
        document_type = "Trading Window Closure"
        event_status = "Compliance notice"
        evidence = _event_evidence(scan, [r"trading window"], chars=600)
        financial_relevance = "Yes"
        financial_relevance_level = "Securities / results-related"
        financial_category = "Trading window / financial results"
        financial_note = "Securities-compliance notice connected with financial results."
        note = "Trading-window notice; no capital/debt infusion."
        reasoning = "This is an insider-trading compliance notice."
    elif re.search(r"regulation\s+74\s*\(?5\)?|certificate pursuant|demat request", descriptor, re.IGNORECASE):
        document_type = "SEBI / RTA Certificate"
        event_status = "Compliance filing"
        evidence = _event_evidence(scan, [r"regulation\s+74\s*\(?5\)?|certificate pursuant|demat request"], chars=600)
        financial_relevance = "Yes"
        financial_relevance_level = "Securities-related"
        financial_category = "Demat / RTA / shareholder securities"
        financial_note = "Share/demat compliance filing; no financing event."
        note = "RTA/SEBI certificate; no capital/debt infusion."
        reasoning = "The filing concerns dematerialisation or registrar compliance."
    elif is_takeover_disclosure:
        document_type = "Shareholding / Takeover Disclosure"
        event_status = "Actual" if re.search(r"acquisition|exceeded\s+5%", scan, re.IGNORECASE) else "Compliance filing"
        material_event = "Yes" if event_status == "Actual" else "No"
        material_event_type = "Significant shareholding change" if event_status == "Actual" else "Shareholding compliance"
        evidence = _event_evidence(scan, [r"exceeded\s+5%|net acquisition|non[- ]encumbrance|takeover regulations|regulation\s+29"], chars=1200)
        key_numbers.extend(extract_amounts(evidence))
        key_numbers.extend(_extract_event_metrics(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Direct securities-related"
        financial_category = "Shareholding / takeover disclosure"
        financial_note = "Securities ownership or takeover-code filing; not automatically a company capital infusion."
        note = "Shareholding/takeover disclosure; financing flags remain off."
        reasoning = "A shareholder acquisition or non-encumbrance filing is tracked separately from company fundraising."
    elif is_notice:
        document_type = "Shareholder / Newspaper Notice"
        event_status = "Notice"
        evidence = _event_evidence(scan, [r"postal ballot|remote e-voting|newspaper publication|newspaper advertisement"], chars=600)
        financial_relevance = "Yes"
        financial_relevance_level = "Shareholder-related"
        financial_category = "Shareholder voting / publication"
        financial_note = "Governance/shareholder notice; no capital/debt infusion."
        note = "Shareholder or newspaper notice."
        reasoning = "This is a notice/publication rather than a financing event."
    elif re.search(r"entered into.{0,180}(?:agreement|memorandum of understanding)|\b(?:agreement|mou)\b.{0,120}(?:with|between)", scan, re.IGNORECASE):
        agreement_patterns = [r"entered into.{0,180}(?:agreement|memorandum of understanding)", r"\b(?:agreement|mou)\b.{0,120}(?:with|between)"]
        document_type = "Commercial Agreement / Partnership"
        event_status = "Actual"
        material_event = "Yes"
        material_event_type = "Commercial agreement / partnership"
        target_counterparty = _extract_target_counterparty(scan)
        evidence = _event_evidence(scan, agreement_patterns, chars=900)
        key_numbers.extend(extract_amounts(evidence))
        key_numbers.extend(_extract_event_metrics(scan))
        financial_relevance = "Yes"
        financial_relevance_level = "Indirect commercial"
        financial_category = "Commercial agreement / partnership"
        financial_note = "Material operating agreement or partnership; not a financing event."
        note = "Commercial agreement/partnership detected."
        reasoning = "The disclosure confirms a business agreement without equity or debt issuance."
    elif is_governance_change:
        document_type = "Management Change"
        event_status = "Actual"
        material_event = "Yes"
        material_event_type = "Management / key personnel change"
        evidence = _event_evidence(scan, [r"appointment|resignation|cessation|change in management"], chars=650)
        key_numbers.extend(_extract_management_change_metrics(scan))
        financial_relevance = "Yes"
        financial_relevance_level = "Indirect governance"
        financial_category = "Management / governance change"
        financial_note = "Material governance event; no capital/debt infusion."
        note = "Management change disclosure."
        reasoning = "The event affects governance rather than financing."
    elif re.search(r"press release|launch(?:es|ed)?|new product|land parcel", descriptor, re.IGNORECASE):
        document_type = "Business Update / Press Release"
        event_status = "Announcement"
        material_event = "Yes"
        material_event_type = "Business / operating update"
        evidence = _event_evidence(scan, [r"press release|launch(?:es|ed)?|new product|land parcel"], chars=750)
        key_numbers.extend(extract_amounts(evidence))
        financial_relevance = "Yes"
        financial_relevance_level = "Indirect commercial"
        financial_category = "Business / operating update"
        financial_note = "Commercial or operating update; no financing event detected."
        note = "Business update/press release."
        reasoning = "The disclosure concerns operations or products, not fundraising."

    key_numbers = _unique_values(key_numbers)
    amount_scan = amount_source or evidence
    evidence_rows_out: list[dict[str, Any]] = []
    if evidence:
        evidence_rows_out.append({
            "Disclosure Date": clean_text(candidate.get("parsed_date")),
            "Company": args.company_name,
            "Symbol": args.symbol,
            "PDF File": pdf_path.name,
            "Category": material_event_type or document_type,
            "Event Status": event_status,
            "Matched Term / Type": document_type,
            "Evidence Snippet": evidence,
            "Amounts Found": "; ".join(_format_labeled_amounts(amount_scan)),
            "Attachment URL": clean_text(candidate.get("actual_pdf_url") or candidate.get("href")),
        })

    amount_rows: list[dict[str, Any]] = []
    amount_seen: set[str] = set()
    for amount, source_start, source_end in _amount_occurrences(amount_scan)[:30]:
        amount_label = _amount_context_label(amount_scan, source_start)
        display_amount = _amount_with_contextual_unit(amount, amount_label, amount_scan)
        page_number: int | str = ""
        amount_context = context_window(amount_scan, source_start, source_end, args.context_chars)
        for page in pages:
            page_text = clean_text(page.get("Text", ""))
            for match in AMOUNT_PATTERN.finditer(page_text):
                page_amount = clean_text(match.group(0))
                if page_amount.lower() != clean_text(amount).lower():
                    continue
                page_label = _amount_context_label(page_text, match.start())
                if amount_label and page_label and amount_label.lower() != page_label.lower():
                    continue
                page_number = page.get("Page", "")
                amount_context = context_window(page_text, match.start(), match.end(), args.context_chars)
                break
            if page_number != "":
                break
        amount_key = "|".join([clean_text(display_amount).lower(), amount_label.lower()])
        if amount_key in amount_seen:
            continue
        amount_seen.add(amount_key)
        amount_rows.append({
            "Disclosure Date": clean_text(candidate.get("parsed_date")),
            "Company": args.company_name,
            "Symbol": args.symbol,
            "PDF File": pdf_path.name,
            "Page": page_number,
            "Amount / Number": display_amount,
            "Amount Type": amount_label,
            "Event Type": material_event_type or document_type,
            "Event Status": event_status,
            "Context": amount_context,
        })

    summary = {
        "Disclosure Date": clean_text(candidate.get("parsed_date")),
        "Company": args.company_name,
        "Symbol": args.symbol,
        "PDF File": pdf_path.name,
        "Subject / Announcement": subject,
        "Financial Relevance?": financial_relevance,
        "Financial Relevance Level": financial_relevance_level,
        "Financial Category": financial_category,
        "Financial Note": financial_note,
        "Material Event?": material_event,
        "Material Event Type": material_event_type,
        "Event Status": event_status,
        "Target / Counterparty": target_counterparty,
        "Document Type": document_type,
        "Infusion Type": infusion_type,
        "Actual Capital Infusion?": capital,
        "Actual Debt Infusion?": debt,
        "Cash / Consideration Disclosed?": cash_disclosed,
        "ESOP Related?": esop_related,
        "ESOP / Stock Options Only?": esop_only,
        "Confidence": confidence,
        "Extracted Result / Note": note,
        "Key Numbers Extracted": "; ".join(key_numbers),
        "Reasoning": reasoning,
        "Evidence Snippet": evidence,
        "Parse Status": parse_status,
        "Content SHA256": file_sha256(pdf_path) if pdf_path.exists() else "",
        "Normalized Text SHA256": normalized_text_sha256(scan),
        "First Page Text SHA256": normalized_text_sha256(pages[0].get("Text", "")) if pages else "",
        "PDF Page Count": len(pages),
        "Downloaded File": str(pdf_path),
        "File Size Bytes": pdf_path.stat().st_size if pdf_path.exists() else "",
        "Announcement Row Context": row_context,
        "Attachment URL": clean_text(candidate.get("actual_pdf_url") or candidate.get("href")),
    }
    return summary, evidence_rows_out, amount_rows


def autosize_worksheet(ws, max_width: int = 70) -> None:
    for column_cells in ws.columns:
        header = column_cells[0]
        letter = header.column_letter
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), max_width))
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), max_width)



def slugify(value: Any, fallback: str = "company") -> str:
    value = clean_text(value)
    value = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return value or fallback


def company_output_prefix(args: argparse.Namespace) -> str:
    return slugify(getattr(args, "symbol", "") or getattr(args, "company_prefix", "") or getattr(args, "company_name", ""), "company").lower()


def is_blank_excel_value(value: Any) -> bool:
    value = clean_text(value)
    return value == "" or value.lower() in {"nan", "none", "null", "nat", "-"}


KNOWN_NSE_SYMBOLS_BY_NAME = {
    # Logistics companies from your current companies.xlsx
    "delhivery ltd": "DELHIVERY",
    "delhivery limited": "DELHIVERY",
    "mahindra logistics ltd": "MAHLOG",
    "mahindra logistics limited": "MAHLOG",
    "allcargo logistics ltd": "ALLCARGO",
    "allcargo logistics limited": "ALLCARGO",
    "transport corporation of india ltd": "TCI",
    "transport corporation of india limited": "TCI",
    "blue dart express ltd": "BLUEDART",
    "blue dart express limited": "BLUEDART",
    "container corporation of india ltd": "CONCOR",
    "container corporation of india limited": "CONCOR",
    "snowman logistics ltd": "SNOWMAN",
    "snowman logistics limited": "SNOWMAN",
    "ceat ltd": "CEATLTD",
    "ceat limited": "CEATLTD",
}


def normalise_company_name_for_lookup(name: str) -> str:
    name = clean_text(name).lower()
    name = name.replace("&", "and")
    name = re.sub(r"[^a-z0-9]+", " ", name).strip()
    name = re.sub(r"\blimited\b", "ltd", name)
    return re.sub(r"\s+", " ", name).strip()


def known_symbol_from_name(company_name: str) -> str:
    raw = clean_text(company_name)
    key = normalise_company_name_for_lookup(raw)
    candidates = {
        key,
        key.replace(" ltd", " limited"),
        key.replace(" limited", " ltd"),
        re.sub(r"\s+", " ", raw.lower().replace(".", "")).strip(),
    }
    for candidate in candidates:
        candidate = normalise_company_name_for_lookup(candidate)
        if candidate in KNOWN_NSE_SYMBOLS_BY_NAME:
            return KNOWN_NSE_SYMBOLS_BY_NAME[candidate]
    return ""


def load_companies_from_excel(excel_path: str | Path) -> list[dict[str, str]]:
    """Read companies from an Excel file.

    Supported columns, case-insensitive:
      - NSE Symbol / Symbol / Ticker
      - Company Name / Company / Name
      - Company Prefix / Prefix

    If only company names are present, the script will try to resolve NSE symbols
    from NSE autocomplete after opening NSE.
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Companies Excel not found: {path}")

    sheets = pd.read_excel(path, sheet_name=None, dtype=str)
    rows: list[dict[str, str]] = []

    def norm_col(c: Any) -> str:
        return re.sub(r"[^a-z0-9]+", "", clean_text(c).lower())

    symbol_keys = {"nsesymbol", "symbol", "ticker", "nse", "nsecode", "nseticker"}
    name_keys = {"companyname", "company", "name", "entity", "entityname"}
    prefix_keys = {"companyprefix", "prefix", "nseprefix", "searchprefix"}

    for sheet_name, df in sheets.items():
        if df is None or df.empty:
            continue

        df = df.dropna(how="all")
        if df.empty:
            continue

        col_map = {norm_col(c): c for c in df.columns}
        symbol_col = next((col_map[k] for k in symbol_keys if k in col_map), None)
        name_col = next((col_map[k] for k in name_keys if k in col_map), None)
        prefix_col = next((col_map[k] for k in prefix_keys if k in col_map), None)

        # If headers are not recognised but there are columns, treat first as company name
        # and second as NSE symbol if available.
        if not symbol_col and not name_col:
            cols = list(df.columns)
            if len(cols) == 1:
                name_col = cols[0]
            elif len(cols) >= 2:
                name_col = cols[0]
                symbol_col = cols[1]

        for _, raw in df.iterrows():
            company_name = clean_text(raw.get(name_col, "")) if name_col else ""
            symbol = clean_text(raw.get(symbol_col, "")) if symbol_col else ""
            prefix = clean_text(raw.get(prefix_col, "")) if prefix_col else ""

            # Ignore blank rows from Excel. Pandas often turns blanks into the string "nan".
            if is_blank_excel_value(company_name):
                company_name = ""
            if is_blank_excel_value(symbol):
                symbol = ""
            if is_blank_excel_value(prefix):
                prefix = ""

            # Ignore rows that are completely blank or look like repeated headers.
            if not company_name and not symbol:
                continue
            if company_name.lower() in {"company name", "company", "name"} and symbol.lower() in {"nse symbol", "symbol", "ticker"}:
                continue

            if not symbol and company_name:
                symbol = known_symbol_from_name(company_name)

            if not company_name and symbol:
                company_name = symbol
            if not prefix:
                prefix = symbol or company_name.split()[0]

            rows.append({
                "company_name": company_name,
                "symbol": symbol.upper().strip(),
                "company_prefix": prefix.upper().strip(),
                "source_sheet": str(sheet_name),
            })

    # Deduplicate by symbol if symbol exists, otherwise company name.
    deduped: list[dict[str, str]] = []
    seen = set()
    for row in rows:
        key = (row.get("symbol") or row.get("company_name", "")).upper().strip()
        if not key or key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped


def deep_find_symbol_items(obj: Any) -> list[dict[str, Any]]:
    """Recursively extract NSE autocomplete-like symbol objects."""
    items = []
    if isinstance(obj, dict):
        if any(k.lower() == "symbol" for k in obj.keys()):
            items.append(obj)
        for value in obj.values():
            items.extend(deep_find_symbol_items(value))
    elif isinstance(obj, list):
        for value in obj:
            items.extend(deep_find_symbol_items(value))
    return items


def resolve_nse_symbol(context, company_name: str, args: argparse.Namespace) -> str:
    """Resolve NSE symbol from company name.

    Order:
      1. Local mapping for common logistics names.
      2. NSE equity master CSV.
      3. NSE autocomplete API as last resort.
    """
    query = clean_text(company_name)
    if not query or is_blank_excel_value(query):
        return ""

    known = known_symbol_from_name(query)
    if known:
        print(f"[NSE] Resolved from local map: '{company_name}' -> {known}")
        return known

    q_norm = normalise_company_name_for_lookup(query)

    csv_urls = [
        "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv",
        "https://archives.nseindia.com/content/equities/EQUITY_L.csv",
        "https://nsearchives.nseindia.com/content/equities/sec_list.csv",
    ]
    for url in csv_urls:
        try:
            response = context.request.get(
                url,
                headers={
                    "Accept": "text/csv,text/plain,*/*",
                    "Referer": NSE_HOME,
                    "User-Agent": args.user_agent,
                },
                timeout=args.download_timeout_seconds * 1000 if args.download_timeout_seconds > 0 else 0,
            )
            if response.status != 200:
                continue

            import csv
            import io
            csv_text = response.text()
            reader = csv.DictReader(io.StringIO(csv_text))
            scored: list[tuple[int, str, str]] = []

            for item in reader:
                symbol = clean_text(item.get("SYMBOL") or item.get("Symbol") or item.get("symbol") or "").upper()
                name = clean_text(item.get("NAME OF COMPANY") or item.get("Security Name") or item.get("SECURITY") or item.get("Company Name") or "")
                if not symbol or not name:
                    continue

                name_norm = normalise_company_name_for_lookup(name)
                score = 0
                if q_norm == name_norm:
                    score += 100
                if q_norm in name_norm or name_norm in q_norm:
                    score += 80
                for part in q_norm.split():
                    if len(part) >= 4 and part in name_norm:
                        score += 10
                if score:
                    scored.append((score, symbol, name))

            scored.sort(reverse=True, key=lambda x: x[0])
            if scored and scored[0][0] >= 50:
                print(f"[NSE] Resolved from NSE CSV: '{company_name}' -> {scored[0][1]} ({scored[0][2]})")
                return scored[0][1]
        except Exception as exc:
            print(f"[NSE] NSE CSV resolver failed for {company_name} via {url}: {exc}")

    autocomplete_urls = [
        f"https://www.nseindia.com/api/search/autocomplete?q={quote_plus(query)}",
        f"https://www.nseindia.com/api/search/autocomplete?keyword={quote_plus(query)}",
    ]
    for url in autocomplete_urls:
        try:
            response = context.request.get(
                url,
                headers={
                    "Accept": "application/json,text/plain,*/*",
                    "Referer": NSE_HOME,
                    "User-Agent": args.user_agent,
                },
                timeout=args.download_timeout_seconds * 1000 if args.download_timeout_seconds > 0 else 0,
            )
            if response.status != 200:
                print(f"[NSE] Symbol resolver HTTP {response.status} for {company_name}")
                continue

            data = response.json()
            items = deep_find_symbol_items(data)
            scored = []
            for item in items:
                symbol = clean_text(item.get("symbol") or item.get("metadata", {}).get("symbol")).upper()
                name = clean_text(item.get("name") or item.get("companyName") or item.get("company") or item.get("description") or "")
                display = normalise_company_name_for_lookup(" ".join([symbol, name]))
                score = 0
                if q_norm == symbol.lower():
                    score += 100
                if q_norm in display:
                    score += 60
                for part in q_norm.split():
                    if len(part) >= 4 and part in display:
                        score += 5
                if symbol:
                    scored.append((score, symbol, name))

            scored.sort(reverse=True, key=lambda x: x[0])
            if scored and scored[0][0] > 0:
                print(f"[NSE] Resolved from autocomplete: '{company_name}' -> {scored[0][1]} ({scored[0][2]})")
                return scored[0][1]
        except Exception as exc:
            print(f"[NSE] Could not resolve NSE symbol for {company_name}: {exc}")

    return ""

def make_run_output_dir(args: argparse.Namespace) -> Path:
    root = Path(args.output_dir).resolve() / datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    root.mkdir(parents=True, exist_ok=True)
    return root


def company_folder_name(index: int, company_args: argparse.Namespace) -> str:
    symbol = slugify(company_args.symbol or company_args.company_name, "company")
    return f"{index:02d}_{symbol}"


def process_one_company(context, page, base_args: argparse.Namespace, company_args: argparse.Namespace, company_output_dir: Path, company_index: int, total_companies: int) -> dict[str, Any]:
    downloads_dir = company_output_dir / "downloaded_disclosures"
    downloads_dir.mkdir(parents=True, exist_ok=True)

    print("\\n" + "=" * 90)
    print(f"[NSE] Company {company_index}/{total_companies}: {company_args.company_name} ({company_args.symbol})")
    print("=" * 90)

    row_summary = {
        "Company Name": company_args.company_name,
        "NSE Symbol": company_args.symbol,
        "Company Prefix": company_args.company_prefix,
        "Company Folder": str(company_output_dir),
        "Status": "Started",
        "PDFs Scanned": 0,
        "Download Failures": 0,
        "Financial Notes": 0,
        "Material Events": 0,
        "Actual Capital Infusion Events": 0,
        "Actual Debt Infusion Events": 0,
        "ESOP / Stock Option Disclosures": 0,
        "Report File": "",
        "Downloaded PDFs Folder": str(downloads_dir),
        "Error": "",
    }

    if is_blank_excel_value(company_args.company_name) and is_blank_excel_value(company_args.symbol):
        row_summary["Status"] = "Skipped"
        row_summary["Error"] = "Blank company row."
        return row_summary

    if not company_args.symbol or is_blank_excel_value(company_args.symbol):
        resolved = resolve_nse_symbol(context, company_args.company_name, base_args)
        if resolved:
            company_args.symbol = resolved
            if not company_args.company_prefix or is_blank_excel_value(company_args.company_prefix):
                company_args.company_prefix = resolved
            row_summary["NSE Symbol"] = company_args.symbol
            row_summary["Company Prefix"] = company_args.company_prefix
        else:
            row_summary["Status"] = "Skipped"
            row_summary["Error"] = "Missing NSE Symbol and could not resolve from company name."
            return row_summary

    try:
        open_page(page, announcements_url(company_args), company_args, f"{company_args.company_name} NSE announcements page")

        if base_args.manual_ready_after_open and (base_args.manual_each_company or company_index == 1):
            print(f"[NSE] Wait until the announcements/PDF list is visible for {company_args.company_name}, then press ENTER.")
            input("Press ENTER to collect disclosure PDFs...")
            wait_for_page_shell_loaded(page, company_args, f"{company_args.company_name} announcements page after manual confirmation")
            wait_until_not_loading(page, company_args, "Disclosure list after manual confirmation")

        candidates = collect_disclosure_candidates(page, company_args)
        print(f"[NSE] Candidate links selected before final PDF dedupe: {len(candidates)}")

        downloaded_rows: list[dict[str, Any]] = []
        result_rows: list[dict[str, Any]] = []
        evidence_rows: list[dict[str, Any]] = []
        amount_rows: list[dict[str, Any]] = []
        processed_pdf_urls: dict[str, str] = {}
        processed_pdf_hashes: dict[str, str] = {}
        processed_text_hashes: dict[str, str] = {}

        for idx, candidate in enumerate(candidates, start=1):
            result = download_candidate(context, candidate, downloads_dir, company_args)
            # Dedupe direct-PDF vs XBRL-wrapper duplicates on the SAME disclosure date,
            # but do not drop a PDF just because the filename/URL lacks the symbol or
            # because NSE reused a similar PDF URL across separate disclosure rows.
            actual_url_key = (result.actual_pdf_url or "").lower().strip()
            actual_key = "|".join([
                actual_url_key,
                clean_text(candidate.get("parsed_date")),
            ]) if actual_url_key else ""

            if actual_key and actual_key in processed_pdf_urls:
                canonical_path = processed_pdf_urls[actual_key]
                if result.downloaded_path and result.downloaded_path.exists():
                    same_as_canonical = False
                    if canonical_path:
                        same_as_canonical = Path(canonical_path).resolve() == result.downloaded_path.resolve()
                    if not same_as_canonical:
                        result.downloaded_path.unlink()
                print(f"[NSE] Skipping duplicate PDF URL for same disclosure date: {result.actual_pdf_url}")
                downloaded_rows.append({
                    "Disclosure Date": clean_text(candidate.get("parsed_date")),
                    "Downloaded?": "No",
                    "Downloaded File": "",
                    "Download Status": "Duplicate URL skipped",
                    "Download Error": "",
                    "Duplicate Of": canonical_path or result.actual_pdf_url,
                    "Content SHA256": "",
                    "Link Text": clean_text(candidate.get("text")),
                    "Row Context": clean_text(candidate.get("context")),
                    "Company Related By Link Text?": "Yes" if candidate.get("companyRelated") else "No",
                    "Original Link URL": clean_text(candidate.get("href")),
                    "Actual PDF URL": result.actual_pdf_url,
                })
                continue

            candidate["actual_pdf_url"] = result.actual_pdf_url
            content_hash = file_sha256(result.downloaded_path) if result.downloaded_path else ""
            content_dedupe_key = disclosure_hash_key(candidate.get("parsed_date"), content_hash)
            duplicate_of = processed_pdf_hashes.get(content_dedupe_key, "") if content_dedupe_key else ""
            if duplicate_of and result.downloaded_path:
                if Path(duplicate_of).resolve() != result.downloaded_path.resolve() and result.downloaded_path.exists():
                    result.downloaded_path.unlink()
                print(f"[NSE] Skipping byte-identical duplicate PDF: {result.downloaded_path.name}")

            canonical_path = duplicate_of or (str(result.downloaded_path) if result.downloaded_path else "")
            if actual_key and canonical_path:
                processed_pdf_urls[actual_key] = canonical_path

            downloaded_rows.append({
                "Disclosure Date": clean_text(candidate.get("parsed_date")),
                "Downloaded?": "No" if duplicate_of else ("Yes" if result.downloaded_path else "No"),
                "Downloaded File": "" if duplicate_of else (str(result.downloaded_path) if result.downloaded_path else ""),
                "Download Status": "Duplicate content skipped" if duplicate_of else result.status,
                "Download Error": result.error,
                "Duplicate Of": duplicate_of,
                "Content SHA256": content_hash,
                "Link Text": clean_text(candidate.get("text")),
                "Row Context": clean_text(candidate.get("context")),
                "Company Related By Link Text?": "Yes" if candidate.get("companyRelated") else "No",
                "Original Link URL": clean_text(candidate.get("href")),
                "Actual PDF URL": result.actual_pdf_url,
            })

            if duplicate_of:
                continue
            if not result.downloaded_path:
                continue

            processed_pdf_hashes[content_dedupe_key] = str(result.downloaded_path)

            summary, evidence, amounts = classify_pdf_result(result.downloaded_path, candidate, company_args)
            download_record = downloaded_rows[-1]
            text_hash = clean_text(summary.get("Normalized Text SHA256"))
            text_dedupe_key = disclosure_hash_key(summary.get("Disclosure Date"), text_hash)
            download_record["Normalized Text SHA256"] = text_hash

            text_duplicate_of = processed_text_hashes.get(text_dedupe_key, "") if text_dedupe_key else ""
            if text_duplicate_of:
                if result.downloaded_path.exists():
                    result.downloaded_path.unlink()
                processed_pdf_hashes.pop(content_dedupe_key, None)
                if actual_key:
                    processed_pdf_urls[actual_key] = text_duplicate_of
                download_record.update({
                    "Downloaded?": "No",
                    "Downloaded File": "",
                    "Download Status": "Duplicate normalized text skipped",
                    "Duplicate Of": text_duplicate_of,
                })
                print(f"[NSE] Skipping text-identical duplicate PDF: {result.downloaded_path.name}")
                continue

            result_rows.append(summary)
            evidence_rows.extend(evidence)
            amount_rows.extend(amounts)
            if text_dedupe_key:
                processed_text_hashes[text_dedupe_key] = str(result.downloaded_path)

        # Final safety pass: ensure every PDF file present in downloaded_disclosures is represented
        # in the Excel, even if its filename/link text did not contain the NSE symbol or it was
        # skipped by a prior candidate rule.
        parsed_paths = {
            str(Path(row.get("Downloaded File", "")).resolve()).lower()
            for row in result_rows
            if clean_text(row.get("Downloaded File"))
        }
        for pdf_file in sorted(downloads_dir.glob("*.pdf")):
            pdf_key = str(pdf_file.resolve()).lower()
            if pdf_key in parsed_paths:
                continue

            fallback_date = parse_disclosure_date(pdf_file.name)
            fallback_date_text = fallback_date.strftime("%Y-%m-%d") if fallback_date else ""
            content_hash = file_sha256(pdf_file)
            content_dedupe_key = disclosure_hash_key(fallback_date_text, content_hash)
            duplicate_of = processed_pdf_hashes.get(content_dedupe_key, "") if content_dedupe_key else ""
            if duplicate_of:
                if Path(duplicate_of).resolve() != pdf_file.resolve() and pdf_file.exists():
                    pdf_file.unlink()
                print(f"[NSE] Safety pass removed byte-identical duplicate PDF: {pdf_file.name}")
                continue
            
            print(f"[NSE] Safety-parse downloaded PDF not yet in results: {pdf_file.name}")
            fallback_candidate = {
                "parsed_date": fallback_date_text,
                "context": pdf_file.name,
                "href": "",
                "actual_pdf_url": "",
                "text": pdf_file.name,
            }
            if content_dedupe_key:
                processed_pdf_hashes[content_dedupe_key] = str(pdf_file)
            downloaded_rows.append({
                "Disclosure Date": fallback_candidate["parsed_date"],
                "Downloaded?": "Yes",
                "Downloaded File": str(pdf_file),
                "Download Status": "Recovered by safety pass",
                "Download Error": "",
                "Duplicate Of": "",
                "Content SHA256": content_hash,
                "Link Text": pdf_file.name,
                "Row Context": pdf_file.name,
                "Company Related By Link Text?": "Unknown",
                "Original Link URL": "",
                "Actual PDF URL": "",
            })
            summary, evidence, amounts = classify_pdf_result(pdf_file, fallback_candidate, company_args)
            text_hash = clean_text(summary.get("Normalized Text SHA256"))
            text_dedupe_key = disclosure_hash_key(summary.get("Disclosure Date"), text_hash)
            downloaded_rows[-1]["Normalized Text SHA256"] = text_hash
            duplicate_path = processed_text_hashes.get(text_dedupe_key, "") if text_dedupe_key else ""
            if duplicate_path:
                if pdf_file.exists():
                    pdf_file.unlink()
                processed_pdf_hashes.pop(content_dedupe_key, None)
                downloaded_rows[-1].update({
                    "Downloaded?": "No",
                    "Downloaded File": "",
                    "Download Status": "Duplicate safety-pass PDF skipped",
                    "Duplicate Of": duplicate_path,
                })
                print(f"[NSE] Safety pass removed semantic duplicate PDF: {pdf_file.name}")
                continue
            result_rows.append(summary)
            evidence_rows.extend(evidence)
            amount_rows.extend(amounts)
            if text_dedupe_key:
                processed_text_hashes[text_dedupe_key] = str(pdf_file)

        print(f"[NSE] Downloaded PDF files in folder: {len(list(downloads_dir.glob('*.pdf')))}")
        print(f"[NSE] PDF result rows written to Excel: {len(result_rows)}")

        report_path = write_report(company_output_dir, downloaded_rows, result_rows, evidence_rows, amount_rows, company_args)
        json_path = company_output_dir / f"{company_output_prefix(company_args)}_financial_notes_FINAL.json"
        json_tmp_path = json_path.with_suffix(".tmp.json")
        json_tmp_path.write_text(
            json.dumps(
                {
                    "downloaded": sanitize_records(downloaded_rows),
                    "results": sanitize_records(result_rows),
                    "evidence": sanitize_records(evidence_rows),
                },
                indent=2,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        json_tmp_path.replace(json_path)

        results_df = pd.DataFrame(result_rows)
        row_summary["Status"] = "Completed"
        row_summary["PDFs Scanned"] = len(result_rows)
        row_summary["Download Failures"] = sum(1 for row in downloaded_rows if row.get("Download Status") == "Failed")
        row_summary["Financial Notes"] = int((results_df.get("Financial Relevance?", pd.Series(dtype=str)) == "Yes").sum()) if not results_df.empty else 0
        row_summary["Material Events"] = int((results_df.get("Material Event?", pd.Series(dtype=str)) == "Yes").sum()) if not results_df.empty else 0
        row_summary["Actual Capital Infusion Events"] = int((results_df.get("Actual Capital Infusion?", pd.Series(dtype=str)) == "Yes").sum()) if not results_df.empty else 0
        row_summary["Actual Debt Infusion Events"] = int((results_df.get("Actual Debt Infusion?", pd.Series(dtype=str)) == "Yes").sum()) if not results_df.empty else 0
        row_summary["ESOP / Stock Option Disclosures"] = int((results_df.get("ESOP / Stock Options Only?", pd.Series(dtype=str)) == "Yes").sum()) if not results_df.empty else 0
        row_summary["Report File"] = str(report_path)
        row_summary["JSON File"] = str(json_path)
        return row_summary

    except Exception as exc:
        row_summary["Status"] = "Failed"
        row_summary["Error"] = f"{type(exc).__name__}: {exc}"
        print(f"[NSE] Company failed: {company_args.company_name} | {row_summary['Error']}")
        return row_summary

def write_report(output_dir: Path, downloaded_rows, result_rows, evidence_rows, amount_rows, args: argparse.Namespace) -> Path:
    """Write and validate the complete workbook before replacing the final file."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / f"{company_output_prefix(args)}_financial_notes_FINAL.xlsx"
    temp_path = report_path.with_name(f"{report_path.stem}.tmp.xlsx")

    downloaded_rows = sanitize_records(list(downloaded_rows))
    result_rows = sanitize_records(list(result_rows))
    evidence_rows = sanitize_records(list(evidence_rows))
    amount_rows = sanitize_records(list(amount_rows))

    results_df = pd.DataFrame(result_rows)
    downloaded_df = pd.DataFrame(downloaded_rows)

    def filtered(column: str, value: str) -> pd.DataFrame:
        if results_df.empty or column not in results_df.columns:
            return pd.DataFrame()
        return results_df[results_df[column].fillna("").astype(str) == value].copy()

    actual_df = pd.DataFrame()
    if not results_df.empty:
        capital_series = results_df.get("Actual Capital Infusion?", pd.Series("No", index=results_df.index))
        debt_series = results_df.get("Actual Debt Infusion?", pd.Series("No", index=results_df.index))
        actual_df = results_df[(capital_series == "Yes") | (debt_series == "Yes")].copy()

    material_df = filtered("Material Event?", "Yes")
    esop_df = filtered("ESOP Related?", "Yes")
    failures_df = pd.DataFrame()
    if not downloaded_df.empty and "Download Status" in downloaded_df.columns:
        failures_df = downloaded_df[downloaded_df["Download Status"] == "Failed"].copy()

    preferred_cols = [
        "Disclosure Date", "Company", "Symbol", "PDF File", "Subject / Announcement",
        "Financial Relevance?", "Financial Relevance Level", "Financial Category", "Financial Note",
        "Material Event?", "Material Event Type", "Event Status", "Target / Counterparty",
        "Document Type", "Infusion Type", "Actual Capital Infusion?", "Actual Debt Infusion?",
        "Cash / Consideration Disclosed?", "ESOP Related?", "ESOP / Stock Options Only?",
        "Confidence", "Extracted Result / Note", "Key Numbers Extracted", "Reasoning",
        "Evidence Snippet", "Parse Status", "Content SHA256", "Downloaded File", "Attachment URL",
    ]
    financial_notes_df = results_df[[col for col in preferred_cols if col in results_df.columns]].copy() if not results_df.empty else pd.DataFrame()

    actual_capital_count = int((results_df.get("Actual Capital Infusion?", pd.Series(dtype=str)) == "Yes").sum()) if not results_df.empty else 0
    actual_debt_count = int((results_df.get("Actual Debt Infusion?", pd.Series(dtype=str)) == "Yes").sum()) if not results_df.empty else 0
    financially_relevant_count = int((results_df.get("Financial Relevance?", pd.Series(dtype=str)) == "Yes").sum()) if not results_df.empty else 0
    material_count = int((results_df.get("Material Event?", pd.Series(dtype=str)) == "Yes").sum()) if not results_df.empty else 0
    review_count = 0
    if not results_df.empty:
        event_review = results_df.get("Event Status", pd.Series("", index=results_df.index)).fillna("").astype(str).eq("Review needed")
        parse_review = results_df.get("Parse Status", pd.Series("", index=results_df.index)).fillna("").astype(str).ne("Parsed")
        review_count = int((event_review | parse_review).sum())
    duplicate_count = int(downloaded_df.get("Download Status", pd.Series(dtype=str)).astype(str).str.startswith("Duplicate").sum()) if not downloaded_df.empty else 0
    successful_download_count = int(downloaded_df.get("Downloaded?", pd.Series(dtype=str)).eq("Yes").sum()) if not downloaded_df.empty else 0

    conclusion = "No completed debt or capital infusion detected."
    if actual_capital_count or actual_debt_count:
        conclusion = f"Completed infusion events detected: {actual_capital_count} capital and {actual_debt_count} debt. See Actual Infusion Events."

    summary_rows = sanitize_records([
        {"Field": "Generated At", "Value": datetime.now().isoformat(timespec="seconds")},
        {"Field": "Company", "Value": args.company_name},
        {"Field": "NSE Symbol", "Value": args.symbol},
        {"Field": "Lookback Days", "Value": args.lookback_days},
        {"Field": "Successfully Downloaded PDFs", "Value": successful_download_count},
        {"Field": "Unique PDFs Scanned", "Value": len(result_rows)},
        {"Field": "Duplicate Candidates Skipped", "Value": duplicate_count},
        {"Field": "Download Failures", "Value": len(failures_df)},
        {"Field": "PDF Rows Listed In Financial Notes", "Value": len(financial_notes_df)},
        {"Field": "Financially Relevant Notes", "Value": financially_relevant_count},
        {"Field": "Material Events", "Value": material_count},
        {"Field": "Actual Capital Infusion Events", "Value": actual_capital_count},
        {"Field": "Actual Debt Infusion Events", "Value": actual_debt_count},
        {"Field": "ESOP-related Disclosures", "Value": len(esop_df)},
        {"Field": "Manual Review Needed", "Value": review_count},
        {"Field": "Conclusion", "Value": conclusion},
        {"Field": "Parser", "Value": "PyMuPDF / fitz with pypdf fallback"},
        {"Field": "Classification Rule", "Value": "Completed-action language required for actual capital/debt flags"},
    ])

    sheets: list[tuple[str, pd.DataFrame, str]] = [
        ("Summary", pd.DataFrame(summary_rows), "No summary available."),
        ("Financial Notes", financial_notes_df, "No PDFs were successfully parsed."),
        ("Material Events", material_df, "No material events detected."),
        ("Actual Infusion Events", actual_df, "No actual capital infusion or debt infusion detected."),
        ("PDF Parsed Results", results_df, "No PDFs were successfully parsed."),
        ("ESOP Stock Options", esop_df, "No ESOP-related disclosures detected."),
        ("Evidence Snippets", pd.DataFrame(evidence_rows), "No event evidence snippets extracted."),
        ("Amounts", pd.DataFrame(amount_rows), "No event-related amounts extracted."),
        ("Downloaded Files", downloaded_df, "No download attempts recorded."),
        ("Download Failures", failures_df, "No download failures."),
    ]

    expected_sheet_names = [name for name, _, _ in sheets]
    if temp_path.exists():
        temp_path.unlink()

    try:
        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            for sheet_name, frame, empty_note in sheets:
                output_frame = frame if not frame.empty else pd.DataFrame([{"Note": empty_note}])
                output_frame.to_excel(writer, index=False, sheet_name=sheet_name)

            workbook = writer.book
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for ws in workbook.worksheets:
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                ws.sheet_view.showGridLines = False
                ws.row_dimensions[1].height = 30
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                autosize_worksheet(ws, max_width=60)
                for col in ws.iter_cols(1, ws.max_column):
                    header = str(col[0].value or "").lower()
                    if any(token in header for token in ["note", "reasoning", "snippet", "context", "url", "file", "subject"]):
                        ws.column_dimensions[col[0].column_letter].width = 60

        check_book = load_workbook(temp_path, read_only=True, data_only=True)
        try:
            if check_book.sheetnames != expected_sheet_names:
                raise RuntimeError(f"Workbook validation failed: sheets are {check_book.sheetnames}")
            if result_rows and check_book["PDF Parsed Results"].max_row != len(result_rows) + 1:
                raise RuntimeError("Workbook validation failed: PDF Parsed Results row count is incomplete")
            if downloaded_rows and check_book["Downloaded Files"].max_row != len(downloaded_rows) + 1:
                raise RuntimeError("Workbook validation failed: Downloaded Files row count is incomplete")
        finally:
            check_book.close()

        temp_path.replace(report_path)
    except Exception:
        if temp_path.exists():
            temp_path.unlink()
        raise

    return report_path


def run(args: argparse.Namespace) -> Path:
    require_dependencies()

    companies_excel = Path(args.companies_excel).resolve()
    companies = load_companies_from_excel(companies_excel)

    if not companies:
        raise RuntimeError(
            f"No companies found in {companies_excel}. "
            "Add at least one row with columns like: Company Name, NSE Symbol, Company Prefix."
        )

    run_output_dir = make_run_output_dir(args)
    print(f"[RUN] Companies Excel: {companies_excel}")
    print(f"[RUN] Companies loaded: {len(companies)}")
    print(f"[RUN] Output folder: {run_output_dir}")

    with sync_playwright() as playwright:
        profile_dir = Path(args.profile_dir).resolve()
        profile_dir.mkdir(parents=True, exist_ok=True)
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=str(profile_dir),
            headless=args.headless,
            viewport={"width": 1440, "height": 1000},
            accept_downloads=True,
            user_agent=args.user_agent,
            args=["--disable-blink-features=AutomationControlled", "--disable-http2"],
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.set_default_timeout(args.default_timeout_seconds * 1000 if args.default_timeout_seconds > 0 else 0)

        try:
            print("[NSE] Opening NSE home once for session cookies.")
            open_page(page, NSE_HOME, args, "NSE home")

            total = len(companies)
            for idx, company in enumerate(companies, start=1):
                company_args = argparse.Namespace(**vars(args))
                company_args.company_name = clean_text(company.get("company_name")) or clean_text(company.get("symbol"))
                company_args.symbol = clean_text(company.get("symbol")).upper()
                company_args.company_prefix = clean_text(company.get("company_prefix")).upper()

                if is_blank_excel_value(company_args.company_name):
                    company_args.company_name = ""
                if is_blank_excel_value(company_args.symbol):
                    company_args.symbol = ""
                if is_blank_excel_value(company_args.company_prefix):
                    company_args.company_prefix = ""

                if not company_args.symbol and company_args.company_name:
                    company_args.symbol = known_symbol_from_name(company_args.company_name)

                if not company_args.company_prefix:
                    company_args.company_prefix = company_args.symbol or (company_args.company_name.split()[0].upper() if company_args.company_name else "")

                # If the company name is actually a symbol and symbol is blank, use it as symbol.
                if not company_args.symbol and company_args.company_name and len(company_args.company_name.split()) == 1:
                    company_args.symbol = company_args.company_name.upper()
                    company_args.company_prefix = company_args.symbol

                folder = run_output_dir / company_folder_name(idx, company_args)
                folder.mkdir(parents=True, exist_ok=True)

                process_one_company(context, page, args, company_args, folder, idx, total)

            print("\\n" + "=" * 90)
            print("Done.")
            print(f"Run output folder: {run_output_dir}")
            print("=" * 90)
            return run_output_dir
        finally:
            if args.keep_browser_open:
                input("Browser left open. Press ENTER to close...")
            context.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Read companies from Excel, download NSE PDFs, and create financial-notes reports per company.")
    parser.add_argument("--companies-excel", default="companies.xlsx", help="Excel file containing companies. Default: companies.xlsx in the current folder.")
    parser.add_argument("--symbol", default=DEFAULT_SYMBOL, help="Fallback single-company NSE symbol if needed.")
    parser.add_argument("--company-prefix", default=DEFAULT_COMPANY_PREFIX, help="Fallback company prefix if needed.")
    parser.add_argument("--company-name", default=DEFAULT_COMPANY_NAME, help="Fallback company name if needed.")
    parser.add_argument("--lookback-days", type=int, default=7, help="How many days back to keep disclosures. Default: 7")
    parser.add_argument("--output-dir", default="outputs", help="Output folder root.")
    parser.add_argument("--profile-dir", default=".nse_browser_profile", help="Persistent browser profile for NSE cookies/session.")
    parser.add_argument("--headless", action="store_true", help="Run without showing browser. Not recommended for NSE.")
    parser.add_argument("--manual-ready", action="store_true", help="Pause after each NSE page open until you press ENTER.")
    parser.add_argument("--manual-ready-after-open", action="store_true", help="Pause after opening first company's announcements page before collecting links.")
    parser.add_argument("--manual-each-company", action="store_true", help="With --manual-ready-after-open, pause for every company, not just the first one.")
    parser.add_argument("--keep-browser-open", action="store_true", help="Keep browser open at the end for debugging.")
    parser.add_argument("--page-is-already-filtered", action="store_true", default=True, help="Kept for compatibility. NSE company pages are treated as already filtered by symbol.")
    parser.add_argument("--require-company-text-in-attachment", action="store_true", help="Old strict mode: only collect links whose filename/text/context contains the symbol/company. Default is OFF.")
    parser.add_argument("--include-undated-visible", action="store_true", help="Include visible disclosure links when a date cannot be parsed from row context.")
    parser.add_argument("--max-documents", type=int, default=0, help="Maximum PDFs to download per company after filtering. 0 means no limit.")
    parser.add_argument("--max-scroll-checks", type=int, default=12, help="Maximum page-scroll/load-more checks per company.")
    parser.add_argument("--stable-rounds", type=int, default=2, help="Stop after this many checks with no new kept links.")
    parser.add_argument("--list-ready-timeout-seconds", type=int, default=180, help="Wait this long for NSE disclosure rows/attachments before asking for manual confirmation.")
    parser.add_argument("--row-update-timeout-seconds", type=int, default=20, help="After scroll/load-more, wait this long for rows to update.")
    parser.add_argument("--scroll-pause-seconds", type=float, default=3.0, help="Pause between row-load checks and scroll actions.")
    parser.add_argument("--after-load-pause-seconds", type=float, default=1.5, help="Extra settling pause after page/list appears loaded.")
    parser.add_argument("--scroll-pixels", type=int, default=300, help="Small scroll distance in pixels. Lower value = slower scrolling.")
    parser.add_argument("--page-timeout-seconds", type=int, default=90, help="Timeout waiting for NSE pages. 0 means no timeout.")
    parser.add_argument("--download-timeout-seconds", type=int, default=90, help="Timeout for download attempts. 0 means no timeout.")
    parser.add_argument("--download-retries", type=int, default=2, help="Retry a failed or incomplete PDF download this many times. Default: 2")
    parser.add_argument("--retry-pause-seconds", type=float, default=2.0, help="Pause between PDF download retries. Default: 2 seconds")
    parser.add_argument("--default-timeout-seconds", type=int, default=30, help="Default Playwright action timeout. 0 means no timeout.")
    parser.add_argument("--context-chars", type=int, default=360, help="Characters of context around each financial keyword/amount.")
    parser.add_argument(
        "--user-agent",
        default=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/136.0.0.0 Safari/537.36"
        ),
        help="Browser user-agent.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    run(parse_args())
